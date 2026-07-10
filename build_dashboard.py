#!/usr/bin/env python3
"""
Build a standalone HTML dashboard visualizing the discrepancy between:

  * Client_Intelligence Factory Anaysis_June11_2026 (1) - v2.xlsx
        -> tab "US Shortfall Analysis"   (demand / "needed" side)
  * AI-Data-Center-Model-CLIENT-April-7-noSKU.xlsx
        -> tab "NA Data Center Supply"   (facility-level supply side)

The central question: does the supply model's US "Planned + UC" capacity match the
"Power Shortfall Before New 'Time to Power' Solutions" (37.71 GW) from the client's
shortfall analysis?

Run:  python build_dashboard.py
Writes: dashboard_data.json  and  dashboard.html  (self-contained, works offline).
"""

import json
import os

import openpyxl
from openpyxl.utils import column_index_from_string as _cidx

HERE = os.path.dirname(os.path.abspath(__file__))

CLIENT_FILE = os.path.join(HERE, "Client_Intelligence Factory Anaysis_June11_2026 (1) - v2.xlsx")
MODEL_FILE = os.path.join(HERE, "AI-Data-Center-Model-CLIENT-April-7-noSKU.xlsx")
NV_FILE = os.path.join(HERE, "Request_NV AI Server Model_052126 2.xlsx")
VENDOR_CHARTJS = os.path.join(HERE, "vendor", "chart.umd.js")
TEMPLATE_FILE = os.path.join(HERE, "dashboard_template.html")
CHIPS_TEMPLATE_FILE = os.path.join(HERE, "chips_template.html")
COMPUTE_TEMPLATE_FILE = os.path.join(HERE, "compute_template.html")
SYNTHESIS_TEMPLATE_FILE = os.path.join(HERE, "synthesis_template.html")
GEO_TEMPLATE_FILE = os.path.join(HERE, "geo_template.html")
FLOWS_TEMPLATE_FILE = os.path.join(HERE, "flows_template.html")
CHIPPOWER_TEMPLATE_FILE = os.path.join(HERE, "chippower_template.html")
DELAYED_TEMPLATE_FILE = os.path.join(HERE, "delayed_template.html")
CANCELLEDPROFILE_TEMPLATE_FILE = os.path.join(HERE, "cancelledprofile_template.html")
SHORTFALLMATH_TEMPLATE_FILE = os.path.join(HERE, "shortfallmath_template.html")

# Offline D3 / map assets (vendored, inlined into the geo + flows pages like Chart.js).
VENDOR_D3 = os.path.join(HERE, "vendor", "d3.min.js")            # full D3 v7 (geo, zoom, force, ...)
VENDOR_SANKEY = os.path.join(HERE, "vendor", "d3-sankey.min.js")
VENDOR_TOPOJSON = os.path.join(HERE, "vendor", "topojson-client.min.js")
VENDOR_US_TOPO = os.path.join(HERE, "vendor", "us-states-10m.json")
VENDOR_PIPELINES = os.path.join(HERE, "vendor", "us-gas-pipelines.json")

OUT_JSON = os.path.join(HERE, "dashboard_data.json")
OUT_HTML = os.path.join(HERE, "dashboard.html")
OUT_CHIPS_HTML = os.path.join(HERE, "chips_dashboard.html")
OUT_COMPUTE_HTML = os.path.join(HERE, "compute_dashboard.html")
OUT_SYNTHESIS_HTML = os.path.join(HERE, "synthesis_dashboard.html")
OUT_GEO_HTML = os.path.join(HERE, "geo_dashboard.html")
OUT_FLOWS_HTML = os.path.join(HERE, "flows_dashboard.html")
OUT_CHIPPOWER_HTML = os.path.join(HERE, "chippower_dashboard.html")
OUT_DELAYED_HTML = os.path.join(HERE, "delayed_dashboard.html")
OUT_CANCELLEDPROFILE_HTML = os.path.join(HERE, "cancelledprofile_dashboard.html")
OUT_SHORTFALLMATH_HTML = os.path.join(HERE, "shortfallmath_dashboard.html")


def col(letter):
    """0-based column index for a spreadsheet column letter."""
    return _cidx(letter) - 1


def num(x):
    return float(x) if isinstance(x, (int, float)) else 0.0


# --------------------------------------------------------------------------- #
# Extraction: US Shortfall Analysis (Client file)
# --------------------------------------------------------------------------- #
def extract_shortfall(wb):
    ws = wb["US Shortfall Analysis"]
    # Random cell access is fine here (tiny tab); pull into a grid first.
    grid = {}
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        for c, v in enumerate(row):
            grid[(r, c)] = v

    def g(letter, row):
        return num(grid.get((row, col(letter))))

    solutions = []
    labels = {
        10: "Nat Gas Turbines",
        11: "Bloom Energy Fuel Cells",
        12: "Site DC at Op'l Nuclear Plant",
        13: "Convert Bitcoin Sites",
    }
    for r, name in labels.items():
        solutions.append({
            "name": name,
            "low": g("C", r),
            "mid": g("D", r),
            "high": g("E", r),
            "prob": g("F", r),
        })

    data = {
        "demand_2026_28": g("C", 3),                 # 67.56
        "less_under_construction": g("C", 4),        # -14.85
        "less_grid_access": g("C", 5),               # -15.0
        "shortfall_before_solutions": g("C", 6),     # 37.71  <-- the "needed" target
        "solutions": solutions,
        "weighted_solutions": {"low": g("C", 15), "mid": g("D", 15), "high": g("E", 15)},
        "net_shortfall": {"low": g("C", 16), "mid": g("D", 16), "high": g("E", 16)},
        "shortfall_pct": {"low": g("C", 17), "mid": g("D", 17), "high": g("E", 17)},
    }
    return data


# --------------------------------------------------------------------------- #
# Extraction: NA Data Center Supply (Model file)
# --------------------------------------------------------------------------- #
YEARS = list(range(2017, 2033))            # 2017..2032
YEAR_COLS = {yr: col(chr(ord("K") + i)) for i, yr in enumerate(YEARS)}  # K..Z


def _quarter_cols():
    """Quarterly capacity columns on 'NA Data Center Supply': AD..BR = 4Q22..4Q32.

    The row-4 header label on each column is the authoritative quarter id (the
    year-end column, e.g. '4Q26' at AT, equals that row's annual 2026 value).
    """
    labs, cols = [], []
    yr, q = 2022, 4
    first, last = col("AD"), col("BR")
    for i in range(last - first + 1):
        labs.append(f"{q}Q{str(yr)[2:]}")
        cols.append(first + i)
        q += 1
        if q > 4:
            q, yr = 1, yr + 1
    return labs, cols


QUARTER_LABELS, QUARTER_COLS = _quarter_cols()


def _norm_type(t):
    if not isinstance(t, str):
        return "other"
    s = t.strip().lower()
    if "hyper" in s:
        return "hyperscaler"
    if "colo" in s:
        return "colocation"
    return "other"


# Tenant classification for the ~10% of US rows tagged with an estimated
# tenant (col DB) or end user (col CZ).  Names normalized to lower-case.
NEOCLOUD_TENANTS = {
    "coreweave", "lambda", "fluidstack", "nebius", "tensorwave", "nscale",
    "iren cloud", "voltage park", "crusoe cloud", "vultr", "togetherai",
    "core42", "cerebras", "bitdeer cloud", "megaspeed", "taiga cloud",
}
HYPERSCALER_TENANTS = {"microsoft", "meta", "google", "aws", "oracle", "apple", "nvidia"}
AI_LAB_USERS = {
    "openai", "openai training", "anthropic", "anthropic training", "gdm",
    "msi", "mai", "xai", "xai inference", "poolside", "bytedance",
}


def _tenant_class(end_user, tenant):
    """Classify a leaf row by its Mosaic-theory tags; None if untagged."""
    t = tenant.strip().lower() if isinstance(tenant, str) else ""
    u = end_user.strip().lower() if isinstance(end_user, str) else ""
    if t in NEOCLOUD_TENANTS:
        return "neocloud_leased"
    if t in HYPERSCALER_TENANTS:
        return "hyperscaler_leased"
    if u in AI_LAB_USERS:
        return "ai_lab_tagged"
    if t or u:
        return "other_tagged"
    return None


# Neocloud names as they appear in the Company column (CV) or GPU-Cloud column
# (DA) — superset of NEOCLOUD_TENANTS, adding the operator spellings used in the
# 'Hyperscalers & Neoclouds' tab.
NEOCLOUD_COMPANIES = NEOCLOUD_TENANTS | {
    "coreweave", "nebius", "nscale", "fluidstack", "iren", "iren cloud",
    "lambda", "tensorwave", "crusoe", "crusoe cloud", "vultr", "togetherai",
    "sharonai", "firmus", "cerebras", "voltage park",
}


def _deploy_class(company, tenant, gpu_cloud, dc_type):
    """Heuristic self-built / leasing / neocloud class for a facility (DERIVED).

    neocloud  -> operated by / leased to a neocloud (company, tenant, or GPU-cloud tag);
    selfbuilt -> a hyperscaler-type facility (owner-operated);
    leasing   -> everything else (colocation space leased by a tenant).
    Cross-checked in aggregate against the 'Hyperscalers & Neoclouds' tab.
    """
    co = company.strip().lower() if isinstance(company, str) else ""
    tn = tenant.strip().lower() if isinstance(tenant, str) else ""
    gc = gpu_cloud.strip().lower() if isinstance(gpu_cloud, str) else ""
    if co in NEOCLOUD_COMPANIES or tn in NEOCLOUD_TENANTS or gc in NEOCLOUD_COMPANIES:
        return "neocloud"
    if _norm_type(dc_type) == "hyperscaler":
        return "selfbuilt"
    return "leasing"


def extract_supply(wb):
    ws = wb["NA Data Center Supply"]

    cA, cBV, cCD, cCE = col("A"), col("BV"), col("CD"), col("CE")
    cType, cCountry, cState, cCompany = col("CX"), col("CT"), col("CP"), col("CV")
    cEndUser, cTenant = col("CZ"), col("DB")
    c2025, c2028 = YEAR_COLS[2025], YEAR_COLS[2028]
    maxcol = max(cA, cBV, cCD, cCE, cType, cCountry, cState, cCompany,
                 cEndUser, cTenant, c2028, *YEAR_COLS.values())

    tot = {"uc": 0.0, "planned": 0.0, "pluc": 0.0, "delta_25_28": 0.0}
    by_type = {"hyperscaler": dict(uc=0.0, planned=0.0, pluc=0.0),
               "colocation": dict(uc=0.0, planned=0.0, pluc=0.0),
               "other": dict(uc=0.0, planned=0.0, pluc=0.0)}
    # per-year capacity (MW) split by type, US-only
    year_by_type = {t: {yr: 0.0 for yr in YEARS} for t in ("hyperscaler", "colocation", "other")}
    by_state, by_company = {}, {}
    tenant_gw = {"neocloud_leased": 0.0, "hyperscaler_leased": 0.0,
                 "ai_lab_tagged": 0.0, "other_tagged": 0.0, "untagged": 0.0}
    n_leaf = 0

    for row in ws.iter_rows(min_row=6, values_only=True):
        if len(row) <= maxcol:
            continue
        a = row[cA]
        # leaf = individual building, identified by a UUID-style token in column A
        if not (isinstance(a, str) and len(a) >= 8 and "-" in a):
            continue
        country = row[cCountry]
        if not (isinstance(country, str) and country.strip().upper() in ("USA", "US", "UNITED STATES")):
            continue

        n_leaf += 1
        uc, planned, pluc = num(row[cBV]), num(row[cCD]), num(row[cCE])
        t = _norm_type(row[cType])
        tot["uc"] += uc
        tot["planned"] += planned
        tot["pluc"] += pluc
        tot["delta_25_28"] += num(row[c2028]) - num(row[c2025])
        by_type[t]["uc"] += uc
        by_type[t]["planned"] += planned
        by_type[t]["pluc"] += pluc
        for yr in YEARS:
            year_by_type[t][yr] += num(row[YEAR_COLS[yr]])

        tc = _tenant_class(row[cEndUser], row[cTenant])
        tenant_gw[tc if tc else "untagged"] += pluc

        st = row[cState]
        if isinstance(st, str) and st.strip():
            by_state[st.strip()] = by_state.get(st.strip(), 0.0) + pluc
        co = row[cCompany]
        if isinstance(co, str) and co.strip():
            by_company[co.strip()] = by_company.get(co.strip(), 0.0) + pluc

    def gw(v):
        return round(v / 1000.0, 3)

    def top(d, n=10):
        items = sorted(d.items(), key=lambda kv: -kv[1])[:n]
        return [{"name": k, "gw": gw(v)} for k, v in items]

    year_total = {yr: sum(year_by_type[t][yr] for t in year_by_type) for yr in YEARS}

    data = {
        "n_leaf": n_leaf,
        "totals_gw": {k: gw(v) for k, v in tot.items()},
        "tenant_gw": {k: gw(v) for k, v in tenant_gw.items()},
        "by_type_gw": {t: {k: gw(v) for k, v in d.items()} for t, d in by_type.items()},
        "years": YEARS,
        "year_series_gw": {
            "hyperscaler": [gw(year_by_type["hyperscaler"][yr]) for yr in YEARS],
            "colocation": [gw(year_by_type["colocation"][yr]) for yr in YEARS],
            "other": [gw(year_by_type["other"][yr]) for yr in YEARS],
            "total": [gw(year_total[yr]) for yr in YEARS],
        },
        "top_states": top(by_state),
        "top_companies": top(by_company),
    }
    return data


def extract_schedule_bridge(wb):
    """Reconcile the US 2025-28 annual-column increase to revised-date capacity.

    The 94.1 GW annual schedule and 47.4 GW revised-date cohort use different
    fields.  This bridge classifies the annual-column increase by revised live
    date (CC), then aligns it to the original 2026-28 cohort (BY) and the phase
    capacity basis (BV + CD) used by the shortfall reconciliation.
    """
    import datetime

    ws = wb["NA Data Center Supply"]
    cA, cCountry = col("A"), col("CT")
    cOrig, cRevised = col("BY"), col("CC")
    cUC, cPlanned = col("BV"), col("CD")
    cCompany, cState = col("CV"), col("CP")
    c25, c28 = YEAR_COLS[2025], YEAR_COLS[2028]
    maxcol = max(cA, cCountry, cOrig, cRevised, cUC, cPlanned,
                 cCompany, cState, c25, c28)

    annual = {"lands_by_2028": 0.0, "slips_past_2028": 0.0,
              "cancelled": 0.0, "no_revised_date": 0.0}
    counts = {k: 0 for k in annual}
    original_lands_annual = 0.0
    original_lands_phase = 0.0
    at_risk_company, at_risk_state = {}, {}

    def yof(v):
        return v.year if isinstance(v, datetime.datetime) else None

    def revised_bucket(ccyr, planned_mw):
        if (ccyr is not None and ccyr >= 2034) or (ccyr is None and planned_mw > 0):
            return "cancelled"
        if ccyr is None:
            return "no_revised_date"
        return "lands_by_2028" if ccyr <= 2028 else "slips_past_2028"

    for row in ws.iter_rows(min_row=6, values_only=True):
        if len(row) <= maxcol:
            continue
        a, country = row[cA], row[cCountry]
        if not (isinstance(a, str) and len(a) >= 8 and "-" in a):
            continue
        if not (isinstance(country, str) and country.strip().upper() in ("USA", "US", "UNITED STATES")):
            continue

        delta = num(row[c28]) - num(row[c25])
        origyr, ccyr = yof(row[cOrig]), yof(row[cRevised])
        planned_mw = num(row[cPlanned])
        bucket = revised_bucket(ccyr, planned_mw)

        if delta > 0:
            annual[bucket] += delta
            counts[bucket] += 1
            if bucket in ("slips_past_2028", "cancelled"):
                co = row[cCompany].strip() if isinstance(row[cCompany], str) else "Unknown"
                st = row[cState].strip() if isinstance(row[cState], str) else "Unknown"
                at_risk_company[co] = at_risk_company.get(co, 0.0) + delta
                at_risk_state[st] = at_risk_state.get(st, 0.0) + delta

        if origyr in (2026, 2027, 2028) and bucket == "lands_by_2028":
            original_lands_annual += max(0.0, delta)
            original_lands_phase += num(row[cUC]) + planned_mw

    annual_total = sum(annual.values())
    outside_original_cohort = annual["lands_by_2028"] - original_lands_annual
    basis_adjustment = original_lands_phase - original_lands_annual
    at_risk = annual["slips_past_2028"] + annual["cancelled"]

    def gw(v):
        return round(v / 1000.0, 3)

    def top(d, n=8):
        return [{"name": k, "gw": gw(v)} for k, v in sorted(d.items(), key=lambda kv: -kv[1])[:n]]

    return {
        "annual_total_gw": gw(annual_total),
        "annual_by_revised_gw": {k: gw(v) for k, v in annual.items()},
        "annual_by_revised_count": counts,
        "annual_lands_by2028_gw": gw(annual["lands_by_2028"]),
        "outside_original_cohort_gw": gw(outside_original_cohort),
        "original_cohort_annual_lands_gw": gw(original_lands_annual),
        "phase_capacity_adjustment_gw": gw(basis_adjustment),
        "revised_cohort_lands_gw": gw(original_lands_phase),
        "at_risk_gw": gw(at_risk),
        "at_risk_pct": round(at_risk / annual_total, 3) if annual_total else None,
        "top_at_risk_companies": top(at_risk_company),
        "top_at_risk_states": top(at_risk_state),
    }


# --------------------------------------------------------------------------- #
# Extraction: geography + deployment (page 5 — map & by-type charts)
# --------------------------------------------------------------------------- #
def extract_geo(wb):
    """Per-facility geography + deployment class for US leaf rows (page-5 map).

    One record per US building whose capacity grows between YE2025 and YE2028
    (i.e. "coming online 2026-2028"); circle size on the map is that GW delta.
    """
    import datetime
    ws = wb["NA Data Center Supply"]
    cA = col("A")
    cLat, cLng, cState, cCity = col("CN"), col("CO"), col("CP"), col("CQ")
    cCountry, cCompany, cType = col("CT"), col("CV"), col("CX")
    cGpu, cTenant, cLive = col("DA"), col("DB"), col("CC")
    cUC = col("BV")
    c25, c28 = YEAR_COLS[2025], YEAR_COLS[2028]
    maxcol = max(cLat, cLng, cState, cCity, cCountry, cCompany, cType,
                 cGpu, cTenant, cLive, cUC, c25, c28)

    facilities = []
    by_class = {"selfbuilt": 0.0, "leasing": 0.0, "neocloud": 0.0}
    for row in ws.iter_rows(min_row=6, values_only=True):
        if len(row) <= maxcol:
            continue
        a = row[cA]
        if not (isinstance(a, str) and len(a) >= 8 and "-" in a):
            continue
        country = row[cCountry]
        if not (isinstance(country, str) and country.strip().upper() in ("USA", "US", "UNITED STATES")):
            continue
        lat, lng = row[cLat], row[cLng]
        if not (isinstance(lat, (int, float)) and isinstance(lng, (int, float))):
            continue
        online = num(row[c28]) - num(row[c25])              # MW online 2026-28
        if online <= 0.5:
            continue
        cls = _deploy_class(row[cCompany], row[cTenant], row[cGpu], row[cType])
        live = row[cLive]
        gw = round(online / 1000.0, 4)
        by_class[cls] += gw
        facilities.append({
            "lat": round(float(lat), 4), "lng": round(float(lng), 4),
            "st": row[cState].strip() if isinstance(row[cState], str) else "",
            "city": row[cCity].strip() if isinstance(row[cCity], str) else "",
            "co": row[cCompany].strip() if isinstance(row[cCompany], str) else "",
            "cls": cls,
            "uc": num(row[cUC]) > 0,                         # under construction vs planned
            "gw": gw,
            "yr": live.year if isinstance(live, datetime.datetime) else None,
        })
    facilities.sort(key=lambda f: -f["gw"])                 # small circles drawn on top
    return {"facilities": facilities, "n": len(facilities),
            "total_gw": round(sum(f["gw"] for f in facilities), 3),
            "by_class_gw": {k: round(v, 3) for k, v in by_class.items()}}


def extract_capacity_by_deploy(wb):
    """US facility capacity (year-end GW) by deployment class, annual + quarterly."""
    ws = wb["NA Data Center Supply"]
    cA, cCountry, cType = col("A"), col("CT"), col("CX")
    cCompany, cGpu, cTenant = col("CV"), col("DA"), col("DB")
    ycols = [YEAR_COLS[y] for y in YEARS]
    maxcol = max(cA, cCountry, cType, cCompany, cGpu, cTenant, max(ycols), max(QUARTER_COLS))
    classes = ("selfbuilt", "leasing", "neocloud")
    annual = {k: [0.0] * len(YEARS) for k in classes}
    quarterly = {k: [0.0] * len(QUARTER_COLS) for k in classes}
    for row in ws.iter_rows(min_row=6, values_only=True):
        if len(row) <= maxcol:
            continue
        a = row[cA]
        if not (isinstance(a, str) and len(a) >= 8 and "-" in a):
            continue
        country = row[cCountry]
        if not (isinstance(country, str) and country.strip().upper() in ("USA", "US", "UNITED STATES")):
            continue
        cls = _deploy_class(row[cCompany], row[cTenant], row[cGpu], row[cType])
        for i, yc in enumerate(ycols):
            annual[cls][i] += num(row[yc])
        for i, qc in enumerate(QUARTER_COLS):
            quarterly[cls][i] += num(row[qc])
    gw = lambda lst: [round(v / 1000.0, 3) for v in lst]
    return {"years": YEARS, "quarters": QUARTER_LABELS,
            "annual": {k: gw(v) for k, v in annual.items()},
            "quarterly": {k: gw(v) for k, v in quarterly.items()}}


def extract_ai_power_us(wb):
    """Client 'AI Power_US': US chip-driven power/energy demand, 2023-2028.

    Rows located by label (robust to drift). Year values sit in cols G/K/O/S/W/AA.
    Returns Data Center Power Capacity (GW), chip server power (GW), and TWh
    (incremental + actual-for-year).
    """
    g = _grid(wb["AI Power_US"], 60, col("AB") + 1)
    years = [2023, 2024, 2025, 2026, 2027, 2028]
    ycols = [col(c) for c in ("G", "K", "O", "S", "W", "AA")]

    def find(pred, scale=1.0, nd=3):
        for r in range(5, 40):
            lab = g.get((r, col("E")))
            if isinstance(lab, str) and pred(lab.strip().lower()):
                return [round(num(g.get((r, c))) * scale, nd) for c in ycols]
        return None

    cap = find(lambda s: s.startswith("data center power capacity"))
    if cap is None:
        raise SystemExit("Could not locate 'Data Center Power Capacity (GW)' in AI Power_US")
    servers_gw = find(lambda s: s.startswith("total power consumption from servers"), scale=1/1000.0)
    twh_incr = find(lambda s: s.startswith("twh - incremental"))
    twh_actual = find(lambda s: s.startswith("twh - actual"))
    return {"years": years, "dc_power_gw": cap, "servers_gw": servers_gw,
            "twh_incr": twh_incr, "twh_actual": twh_actual}


# Vendor families for the global chip-power breakdown (order = keyword priority).
def _chip_family(gen):
    g = gen.lower()
    if any(k in g for k in ("china", "ascend", "maia", "bytedance", "meta", "openai", "microsoft", "xai")):
        return "Other / custom"
    if "trainium" in g or "inferentia" in g:
        return "AWS"
    if "tpu" in g or "ironwood" in g or "mediatek" in g or g[:2] in ("v7", "v8") or g.startswith("n+"):
        return "Google TPU"
    if any(k in g for k in ("gaudi", "falcon", "intel")):
        return "Intel"
    if g.startswith("mi") or "amd" in g:
        return "AMD"
    if any(k in g for k in ("h100", "a100", "b100", "r100", "r+", "other gpu", "nvda", "nvidia")):
        return "Nvidia"
    return "Other / custom"


FAMILY_ORDER = ["Nvidia", "AMD", "Intel", "Google TPU", "AWS", "Other / custom"]


def extract_chip_power_global(wb):
    """Client 'AI Power_Global': GW of chip power by vendor family (server power),
    global DC power capacity (GW, incl PUE) and TWh, 2021-2028.

    The tab computes a per-generation power block (volume -> servers -> Max Input
    Power (W) -> MW). We sum each generation's 'Total Server Power Consumption (MW)'
    into vendor families, and read the tab's own summary GW/TWh rows.
    """
    ws = wb["AI Power_Global"]
    rows = list(ws.iter_rows(min_row=1, max_row=422, values_only=True))
    cB, cC = col("B"), col("C")
    ycols = [col(c) for c in ("H", "I", "J", "K", "L", "M", "N", "O")]   # 2021..2028
    years = list(range(2021, 2029))

    def lbl(row):
        for c in (cC, cB):
            v = row[c] if c < len(row) else None
            if isinstance(v, str) and v.strip():
                return v.strip()
        return ""

    def vals(row, scale=1.0, nd=3):
        return [round(num(row[i]) * scale, nd) if i < len(row) else 0.0 for i in ycols]

    fam = {f: [0.0] * len(years) for f in FAMILY_ORDER}
    cur = None
    for row in rows:
        L = lbl(row)
        if L.endswith("Power Usage (MW)"):
            cur = L[: -len("Power Usage (MW)")].strip()
            continue
        if cur and "Total Server Power Consumption (MW)" in L:
            f = _chip_family(cur)
            v = vals(row, scale=1 / 1000.0)          # MW -> GW
            fam[f] = [a + b for a, b in zip(fam[f], v)]
            cur = None

    dc_power_gw = twh_incr = twh_actual = None
    for row in rows:
        L = lbl(row)
        if L == "Data Center Power Capacity (GW)" and dc_power_gw is None:
            dc_power_gw = vals(row)
        elif L.startswith("TWh - Incremental"):
            twh_incr = vals(row, nd=1)
        elif L.startswith("TWh - Actual"):
            twh_actual = vals(row, nd=1)
    if dc_power_gw is None:
        raise SystemExit("Could not locate global 'Data Center Power Capacity (GW)' in AI Power_Global")

    families = [{"name": f, "values": [round(x, 3) for x in fam[f]]} for f in FAMILY_ORDER]
    total_servers_gw = [round(sum(fam[f][i] for f in FAMILY_ORDER), 3) for i in range(len(years))]
    return {"years": years, "families": families, "total_servers_gw": total_servers_gw,
            "dc_power_gw": dc_power_gw, "twh_incr": twh_incr, "twh_actual": twh_actual}


def derive_chips_by_deploy(ai_power_us, buildout):
    """Split the client US GW-of-chips total by the model's deployment mix.

    Magnitude: client 'AI Power_US' Data Center Power Capacity (GW).
    Mix %: 'Hyperscalers & Neoclouds' deployment types (self-build / leasing /
    neocloud share per year). This is a DERIVED cross-source blend.
    """
    hn_years = buildout["years"]                            # 2017..2032
    n = len(hn_years)
    sb = [0.0] * n; ls = [0.0] * n; ne = [0.0] * n
    for h in buildout["hyperscalers"]:
        for k, s in h["sub"].items():
            kl = k.lower()
            if "self-build" in kl:
                sb = [a + b for a, b in zip(sb, s)]
            elif "leasing" in kl:
                ls = [a + b for a, b in zip(ls, s)]
            elif "contracted neocloud" in kl:
                ne = [a + b for a, b in zip(ne, s)]
    for nc in buildout["neoclouds"]:                        # whole neocloud company = neocloud
        for k, s in nc["sub"].items():
            if "self-build" in k.lower() or "leasing" in k.lower():
                ne = [a + b for a, b in zip(ne, s)]

    years = ai_power_us["years"]
    tot = ai_power_us["dc_power_gw"]
    out = {"years": years, "selfbuilt": [], "leasing": [], "neocloud": [],
           "mix_pct": {"selfbuilt": [], "leasing": [], "neocloud": []}}
    for i, y in enumerate(years):
        j = hn_years.index(y)
        denom = sb[j] + ls[j] + ne[j]
        fs, fl, fn = (sb[j] / denom, ls[j] / denom, ne[j] / denom) if denom > 0 else (0.0, 0.0, 0.0)
        out["selfbuilt"].append(round(tot[i] * fs, 3))
        out["leasing"].append(round(tot[i] * fl, 3))
        out["neocloud"].append(round(tot[i] * fn, 3))
        out["mix_pct"]["selfbuilt"].append(round(fs, 4))
        out["mix_pct"]["leasing"].append(round(fl, 4))
        out["mix_pct"]["neocloud"].append(round(fn, 4))
    return out


# --------------------------------------------------------------------------- #
# Natural-gas pipeline proximity (vendored EIA pipeline geometry)
# --------------------------------------------------------------------------- #
_EARTH_MI = 3958.8
_PIPE_CELL = 0.5            # grid cell size in degrees for the spatial index
_PIPE_MAXCELL = 2          # search +/- this many cells (~1 deg / up to ~50-69 mi)


def load_pipelines():
    """Vendored EIA gas-pipeline geometry: returns list of [lng,lat] polylines."""
    with open(VENDOR_PIPELINES, "r", encoding="utf-8") as f:
        gj = json.load(f)
    return gj["coordinates"]


def _seg_dist_mi(plng, plat, a, b):
    """Great-circle-ish distance (mi) from point to segment a-b via a local
    equirectangular projection centered on the point (accurate at these scales)."""
    import math
    cosl = math.cos(math.radians(plat))
    ax = math.radians(a[0] - plng) * cosl * _EARTH_MI
    ay = math.radians(a[1] - plat) * _EARTH_MI
    bx = math.radians(b[0] - plng) * cosl * _EARTH_MI
    by = math.radians(b[1] - plat) * _EARTH_MI
    dx, dy = bx - ax, by - ay
    if dx == 0 and dy == 0:
        return math.hypot(ax, ay)
    t = max(0.0, min(1.0, -(ax * dx + ay * dy) / (dx * dx + dy * dy)))
    return math.hypot(ax + t * dx, ay + t * dy)


def annotate_pipe_proximity(geo, segs):
    """Add `pipe_mi` (miles to nearest gas pipeline) to each facility; add a
    by-bucket GW summary to `geo`. Uses a lat/long grid index over pipeline
    vertices so each facility only tests nearby segments."""
    import math
    grid = {}
    for i, seg in enumerate(segs):
        for x, y in seg:
            grid.setdefault((math.floor(x / _PIPE_CELL), math.floor(y / _PIPE_CELL)), set()).add(i)

    def nearest_mi(plng, plat):
        cx, cy = math.floor(plng / _PIPE_CELL), math.floor(plat / _PIPE_CELL)
        cand = set()
        for a in range(-_PIPE_MAXCELL, _PIPE_MAXCELL + 1):
            for b in range(-_PIPE_MAXCELL, _PIPE_MAXCELL + 1):
                cand |= grid.get((cx + a, cy + b), set())
        best = 1e9
        for i in cand:
            s = segs[i]
            for k in range(len(s) - 1):
                d = _seg_dist_mi(plng, plat, s[k], s[k + 1])
                if d < best:
                    best = d
        return round(best, 2) if best < 1e8 else None

    buckets = {"le5": 0.0, "le25": 0.0, "gt25": 0.0}   # GW within <=5 / 5-25 / >25 mi
    for f in geo["facilities"]:
        d = nearest_mi(f["lng"], f["lat"])
        f["pipe_mi"] = d
        if d is None or d > 25:
            buckets["gt25"] += f["gw"]
        elif d <= 5:
            buckets["le5"] += f["gw"]
        else:
            buckets["le25"] += f["gw"]
    geo["pipe_buckets_gw"] = {k: round(v, 3) for k, v in buckets.items()}
    return geo


# --------------------------------------------------------------------------- #
# Extraction: chips & buildout (all three workbooks)
# --------------------------------------------------------------------------- #
def _grid(ws, max_row, max_col=60):
    """Read a bounded rectangle into a {(row, col0): value} dict (1-based rows)."""
    g = {}
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=True), start=1):
        for c, v in enumerate(row[:max_col]):
            if v is not None:
                g[(r, c)] = v
    return g


def _parse_watts(s):
    """'1,400W' -> 1400 ; returns None for multi-value or non-numeric cells."""
    if isinstance(s, (int, float)):
        return float(s)
    if not isinstance(s, str) or "|" in s or "\n" in s:
        return None
    digits = "".join(ch for ch in s if ch.isdigit())
    return float(digits) if digits else None


def extract_nv(path):
    """Morgan Stanley Nvidia AI-server model: chips, racks, customers, TDP, rack kW."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # -- ODM share analysis: chip families (rows 2-7) + customer rack blocks --
    g = _grid(wb["ODM share analysis"], 220, 10)
    chip_years = [int(g[(1, c)]) for c in range(col("D"), col("H") + 1)]   # 2023..2027
    chip_families = []
    for r in range(2, 8):
        name = g.get((r, col("B")))
        if not isinstance(name, str):
            continue
        chip_families.append({
            "name": name.strip(),
            "values": [round(num(g.get((r, c))), 1) for c in range(col("D"), col("H") + 1)],
        })

    def customer_block(r_first, r_last):
        out = []
        for r in range(r_first, r_last + 1):
            name = g.get((r, col("B")))
            if not isinstance(name, str):
                continue
            name = name.strip()
            if name in ("Mix", "Other non-hyperscalers"):   # subtotal of the rows below it
                continue
            out.append({"name": name, "kracks": round(num(g.get((r, col("D")))), 3)})
        return out

    customers = {
        "2025": customer_block(171, 183),
        "2026": customer_block(203, 215),
        "total_2025": round(num(g.get((170, col("D")))), 3),
        "total_2026": round(num(g.get((202, col("D")))), 3),
    }

    # -- GB200-300 Tracker: monthly ramp + annual racks by SKU --
    g = _grid(wb["GB200-300 Tracker"], 60, 62)
    months, monthly = [], []
    yr = ""
    for c in range(col("C"), col("Z") + 1):
        lab = g.get((39, c))
        if not isinstance(lab, str):
            continue
        if c == col("C"):
            yr = "25"
        if c == col("O"):
            yr = "26"
        months.append(f"{lab} '{yr}")
        monthly.append(round(num(g.get((40, c))), 0))

    rack_years = [int(g[(39, c)]) for c in (col("BE"), col("BF"), col("BG"))]  # 2025..2027
    rack_sku_rows = [(40, "GB200 NVL72"), (41, "GB300 NVL72"), (42, "VR200 NVL72"), (43, "VR300")]
    rack_by_sku = [{
        "name": name,
        "values": [round(num(g.get((r, c))), 0) for c in (col("BE"), col("BF"), col("BG"))],
    } for r, name in rack_sku_rows]
    rack_totals = [round(num(g.get((44, c))), 0) for c in (col("BE"), col("BF"), col("BG"))]

    # -- GPU Roadmap: TDP per SKU (row 8) --
    g = _grid(wb["GPU Roadmap"], 10, 14)
    ms_tdp = []
    for c in range(col("C"), col("M") + 1):
        sku = g.get((3, c))
        w = _parse_watts(g.get((8, c)))
        if isinstance(sku, str) and w:
            ms_tdp.append({"sku": sku.split("\n")[0].strip(), "watts": w})

    # -- Nvidia Rack Layout: per-rack power usage kW (row 55) --
    g = _grid(wb["Nvidia Rack Layout"], 56, 14)
    rack_kw = {
        "GB200 NVL72": round(num(g.get((55, col("C")))), 2),
        "GB300 NVL72": round(num(g.get((55, col("E")))), 2),
        "VR200 NVL72": round(num(g.get((55, col("G")))), 2),
    }

    wb.close()
    return {
        "chip_years": chip_years, "chip_families": chip_families,
        "customers": customers,
        "months": months, "monthly_racks": monthly,
        "rack_years": rack_years, "rack_by_sku": rack_by_sku, "rack_totals": rack_totals,
        "ms_tdp": ms_tdp, "rack_kw": rack_kw,
    }


def extract_chip_watts(wb):
    """Client model 'Power Requirements per Chip': watts per SKU (unit col H == 'W').

    The tab lists ~150 SKUs across three repeated blocks (Nvidia/TPU/Trainium/Maia/
    AMD/boutique chips); a prior version of this function stopped at row 50 and
    silently dropped everything past the Nvidia+TPU section (all of Microsoft Maia,
    AMD's MI-series, AWS Inferentia3/Trainium4, and several boutique chips). Scan
    the full sheet and keep only the first occurrence of each SKU name so the
    later repeated blocks (some are '%' assumptions, not watts) don't overwrite it.
    """
    ws = wb["Power Requirements per Chip"]
    cE, cH, cK, cW = col("E"), col("H"), col("K"), col("W")
    seen = set()
    out = []
    for row in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=True):
        if len(row) <= cW:
            continue
        name, unit = row[cE], row[cH]
        if not (isinstance(name, str) and isinstance(unit, str) and unit.strip() == "W"):
            continue
        key = name.strip()
        if key in seen:
            continue
        watts = max((num(row[c]) for c in range(cK, cW + 1)), default=0.0)
        if watts > 0:
            seen.add(key)
            out.append({"sku": key, "watts": watts})
    return out


HN_YEARS = list(range(2017, 2033))  # 'Hyperscalers & Neoclouds' year cols K..Z (row 4)


def extract_buildout(wb):
    """'Hyperscalers & Neoclouds' company blocks + AI-lab MW from 'AI CPU Demand calculations'."""
    ws = wb["Hyperscalers & Neoclouds"]
    cB, cC, cD = col("B"), col("C"), col("D")
    ycols = [YEAR_COLS[yr] for yr in HN_YEARS]  # same K..Z layout as the supply tab

    def series(row):
        return [round(num(row[c]), 1) for c in ycols]

    hyperscalers, neoclouds = [], []
    global_neocloud = None
    block = hyperscalers
    current = None
    for row in ws.iter_rows(min_row=5, max_row=140, values_only=True):
        if len(row) <= max(ycols):
            continue
        b, c, d = row[cB], row[cC], row[cD]
        if isinstance(b, str) and "Global Neocloud" in b:
            block = neoclouds
            global_neocloud = series(row)
            current = None
            continue
        if isinstance(c, str) and c.strip() not in ("", "company"):
            name = c.strip()
            if "Neocloud capacity" in name:      # cross-reference row, not a company
                current = None
                continue
            current = {"name": name, "total": series(row), "sub": {}}
            block.append(current)
            continue
        if current is not None and isinstance(d, str) and d.strip():
            current["sub"][d.strip()] = series(row)

    # AI labs: rows with 'Identified Datacenters' in col C
    ws = wb["AI CPU Demand calculations"]
    ai_labs = []
    for row in ws.iter_rows(min_row=40, max_row=70, values_only=True):
        if len(row) <= max(ycols):
            continue
        c = row[cC]
        if isinstance(c, str) and "Identified Datacenter" in c:
            name = c.replace("Identified Datacenters", "").replace("Identified Datacenter", "").strip()
            ai_labs.append({"name": name, "values": series(row)})

    return {"years": HN_YEARS, "hyperscalers": hyperscalers, "neoclouds": neoclouds,
            "global_neocloud": global_neocloud, "ai_labs": ai_labs}


def extract_client_chips(wb):
    """Client file: global GPU/ASIC volumes, US DC power (Part IIB), ExaFLOPs need vs supply,
    TWh series, PUE/utilization, per-SKU volume rows, ExaFLOPs by generation, training wall."""
    g = _grid(wb["Power Summary"], 100, 20)
    years = list(range(2023, 2029))                      # cols D..I and L..Q

    def brow(r, c1="D", c2="I"):
        return [round(num(g.get((r, c))), 1) for c in range(col(c1), col(c2) + 1)]

    total_gpus = brow(43)
    total_asics = brow(63)
    # r57 'China/Ascend/MAIA 100/Other' is a grab-bag (Huawei Ascend, China parts,
    # MAIA, unspecified "Other"). Strip it so the ASIC line is comparable merchant
    # silicon (Google TPU, AWS Trainium, Meta/OpenAI/MSFT/xAI ASICs).
    asics_china_ascend_other = brow(57)
    total_asics_ex_china = [round(a - b, 1) for a, b in zip(total_asics, asics_china_ascend_other)]
    us_power_gw = brow(46, "L", "Q")
    us_additions_gw = brow(47, "L", "Q")

    # -- TWh rows (Part IIA global / IIB US) --
    twh_genai_global = brow(10, "L", "Q")            # 27.2 -> 801.6
    twh_base_global = brow(14, "L", "Q")             # 341 -> 549
    twh_genai_us = brow(39, "L", "Q")                # 16.3 -> 455.1
    global_dc_gw = brow(18, "L", "Q")                # 56 -> 205.6

    # -- PUE / utilization located by label (fixed addresses drift) --
    pue, util = None, None
    for r in range(1, 30):
        b = g.get((r, col("B")))
        if isinstance(b, str):
            if "Data Center PUE" in b:
                pue = [round(num(g.get((r, c))), 3) for c in range(col("D"), col("I") + 1)]
            elif "GPU Server Utilization Rate" in b:
                util = num(g.get((r, col("C"))))
    if pue is None or not util:
        raise SystemExit("Could not locate PUE / utilization rows by label in Power Summary")

    # -- per-SKU volume rows (labels col B) for the derived energy split --
    def vol_rows(r1, r2, skip=()):
        out = []
        for r in range(r1, r2 + 1):
            b = g.get((r, col("B")))
            if isinstance(b, str) and r not in skip:
                out.append({"row": r, "label": b.strip(), "values": brow(r)})
        return out

    gpu_vol_rows = vol_rows(24, 42)
    asic_vol_rows = vol_rows(45, 62, skip=(57,))     # China/Ascend/MAIA/Other excluded

    # -- ExaFLOPs by generation (rows 81-85) --
    exa_by_gen = []
    for r in range(81, 86):
        b = g.get((r, col("B")))
        if isinstance(b, str):
            exa_by_gen.append({"name": b.strip().strip('"'), "values": brow(r, "E", "I")})
    exa_total = brow(90, "E", "I")

    g = _grid(wb["Compute Demand"], 120, 10)
    exa_years = [2026, 2027, 2028]

    def crow(r, c1="G", c2="I", d=1):
        return [round(num(g.get((r, c))), d) for c in range(col(c1), col(c2) + 1)]

    return {
        "years": years,
        "total_gpus": total_gpus, "total_asics": total_asics,
        "total_asics_ex_china": total_asics_ex_china,
        "asics_china_ascend_other": asics_china_ascend_other,
        "us_power_gw": us_power_gw, "us_additions_gw": us_additions_gw,
        "twh_genai_global": twh_genai_global, "twh_base_global": twh_base_global,
        "twh_genai_us": twh_genai_us, "global_dc_gw": global_dc_gw,
        "pue": pue, "utilization": util,
        "gpu_vol_rows": gpu_vol_rows, "asic_vol_rows": asic_vol_rows,
        "exa_gen_years": [2024, 2025, 2026, 2027, 2028],
        "exa_by_gen": exa_by_gen, "exa_total": exa_total,
        "exa_years": exa_years,
        "exa_cumulative": crow(40),
        "exa_needed": crow(44),
        "exa_shortage": crow(45),
        "exa_shortage_pct": crow(46, d=4),
        # training wall (years E..I = 2026-2030 for r104; E..G = 2026-28 for r110/111)
        "training_years": [2026, 2027, 2028, 2029, 2030],
        "training_power_mw": crow(104, "E", "I"),
        "training_share_years": [2026, 2027, 2028],
        "training_share": crow(111, "E", "G", d=4),
        "bw_equiv_sold": crow(110, "E", "G"),
    }


# --------------------------------------------------------------------------- #
# Extraction: pages 3/4 — customer demand, unconstrained, balance, revisions,
# slippage, per-company additions
# --------------------------------------------------------------------------- #
def extract_customer_demand(wb):
    """'AI Demand by Customer': constrained demand aggregates + per-company AI power."""
    g = _grid(wb["AI Demand by Customer"], 300, 30)
    years = list(range(2017, 2033))                          # K..Z, header row 4
    ycols = [YEAR_COLS[yr] for yr in years]

    def series(r):
        return [round(num(g.get((r, c))), 1) for c in ycols]

    companies = [{"name": n, "values": series(r)} for r, n in
                 [(223, "Microsoft"), (224, "Meta"), (225, "Google"),
                  (226, "Amazon"), (227, "Oracle"), (228, "Apple")]]
    # neocloud/China blocks only carry 2023-2026 (cols Q..T)
    def qt(r):
        return [round(num(g.get((r, c))), 1) for c in (col("Q"), col("R"), col("S"), col("T"))]

    return {
        "years": years,
        "total_units_k": series(62),
        "install_base_k": series(117),
        "ai_power_mw": series(177),           # constrained cumulative AI critical IT power
        "hyperscaler_ai_mw": series(229),
        "companies": companies,
        "type_years": [2023, 2024, 2025, 2026],
        "type_hyperscaler": [round(num(g.get((229, c))), 1) for c in (col("Q"), col("R"), col("S"), col("T"))],
        "type_neoclouds": qt(258),
        "type_china": qt(290),
        "type_total": [round(num(g.get((177, c))), 1) for c in (col("Q"), col("R"), col("S"), col("T"))],
        "neo_coreweave": qt(249), "neo_tesla": qt(252), "neo_xai": qt(255),
    }


def extract_unconstrained(wb):
    """'Unconstrained AI Demand': chip-supply-driven demand curve (years J=2014..Z=2030)."""
    g = _grid(wb["Unconstrained AI Demand"], 370, 30)
    years = list(range(2014, 2031))
    ycols = [col("J") + i for i in range(len(years))]
    return {
        "years": years,
        "ai_power_mw": [round(num(g.get((365, c))), 1) for c in ycols],
        "total_units_k": [round(num(g.get((124, c))), 1) for c in ycols],
    }


def extract_na_balance(wb):
    """'NA DC Supply-Demand': the model's own supply-vs-demand balance (K=2017..X=2030)."""
    g = _grid(wb["NA DC Supply-Demand"], 160, 30)
    years = list(range(2017, 2031))
    ycols = [col("K") + i for i in range(len(years))]

    def series(r):
        return [round(num(g.get((r, c))), 1) for c in ycols]

    return {
        "years": years,
        "incr_demand_mw": series(154),
        "capacity_added_mw": series(155),
        "surplus_deficit_mw": series(156),
        "cum_surplus_deficit_mw": series(108),
        "total_demand_mw": series(100),
        "total_capacity_mw": series(91),
    }


def extract_revisions(wb):
    """'Revisions vs prior versions': Total Global Capacity vintage ladder (rows 11-51)
    + Oracle Overseas % revision row (r844)."""
    ws = wb["Revisions vs prior versions"]
    vintages = []
    oracle_pct = None
    for r, row in enumerate(ws.iter_rows(min_row=11, max_row=850, values_only=True), start=11):
        if len(row) <= col("X"):
            continue
        gg = row[col("G")]
        if r <= 51 and isinstance(gg, str) and gg.strip() != "% revision":
            vintages.append({
                "label": gg.strip(),
                "y2025": round(num(row[col("S")]), 0),
                "y2026": round(num(row[col("T")]), 0),
                "y2027": round(num(row[col("U")]), 0),
                "y2028": round(num(row[col("V")]), 0),
                "y2030": round(num(row[col("X")]), 0),
            })
            if gg.strip() == "First version":
                pass                                      # last row of the block
        if r == 844:                                       # Oracle Overseas '% revision'
            oracle_pct = [round(num(row[c]), 3) for c in
                          (col("T"), col("U"), col("V"), col("W"), col("X"))]
    vintages.reverse()                                     # oldest -> newest for charting
    return {"vintages": vintages,
            "oracle_pct_years": [2026, 2027, 2028, 2029, 2030],
            "oracle_pct": oracle_pct}


def extract_slippage(wb):
    """'NA Data Center Supply': modeled go-live slippage (Actual Live minus nameplate), US leaf rows."""
    import datetime
    ws = wb["NA Data Center Supply"]
    cA, cBY, cCB, cCT = col("A"), col("BY"), col("CB"), col("CT")
    diffs = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if len(row) <= max(cA, cBY, cCB, cCT):
            continue
        a = row[cA]
        if not (isinstance(a, str) and len(a) >= 8 and "-" in a):
            continue
        ct = row[cCT]
        if not (isinstance(ct, str) and ct.strip().upper() == "USA"):
            continue
        by, cb = row[cBY], row[cCB]
        if isinstance(by, datetime.datetime) and isinstance(cb, datetime.datetime):
            diffs.append((cb - by).days / 30.44)
    diffs.sort()
    n = len(diffs)
    buckets = {"~0": 0, "1-3 mo": 0, "4-6 mo": 0, "7-12 mo": 0, ">12 mo": 0}
    for d in diffs:
        if d <= 0.5:
            buckets["~0"] += 1
        elif d <= 3.5:
            buckets["1-3 mo"] += 1
        elif d <= 6.5:
            buckets["4-6 mo"] += 1
        elif d <= 12.5:
            buckets["7-12 mo"] += 1
        else:
            buckets[">12 mo"] += 1
    return {"n": n,
            "median_mo": round(diffs[n // 2], 1) if n else None,
            "mean_mo": round(sum(diffs) / n, 1) if n else None,
            "buckets": buckets,
            "pct_ge_7mo": round(100 * (buckets["7-12 mo"] + buckets[">12 mo"]) / n, 1) if n else None}


def extract_dashboard_adds(wb):
    """Model 'Dashboard' rows 17-25: per-company annual capacity additions (MW/yr, 2018-2032)."""
    g = _grid(wb["Dashboard"], 30, 30)
    years = list(range(2018, 2033))                        # L..Z
    ycols = [col("L") + i for i in range(len(years))]
    companies = []
    for r in range(17, 26):
        name = g.get((r, col("C")))
        if isinstance(name, str):
            companies.append({"name": name.strip(),
                              "values": [round(num(g.get((r, c))), 1) for c in ycols]})
    return {"years": years, "companies": companies}


# --------------------------------------------------------------------------- #
# Cancelled & delayed sites (page 8)
# --------------------------------------------------------------------------- #
# The supply model carries NO status flag. "Cancelled/delayed" is therefore
# DERIVED two ways, kept side by side:
#   (a) in-model: each US facility's "Live Date - Quarter End" (col CC). Sites
#       dated 2029-2033 are "delayed past 2028"; a far-future sentinel (>=2034,
#       e.g. 2043/2099) is treated as "cancelled / indefinite".
#   (b) external, CITED: specific projects publicly reported cancelled/delayed,
#       with a source link. Named-site delay dates are rarely reported, so the
#       delayed *magnitude* comes from (a); these citations name/ground it.
# Every figure the user reads on this page is one or the other, both labeled.
#
# Sources verified/retrieved during the research pass (July 2026). Confidence is
# noted; a few outlets (CNBC, Data Center Watch) 403'd and are corroborated via
# the outlets listed. "match" is the Company-column value in the supply model so
# the page can attach the model's own delayed+cancelled GW for that operator.
# Externally-reported, individually-cited US projects (originals broadly slated
# 2026-2028) that were cancelled or delayed. Coordinates are approximate city/
# county centroids (labelled "approx" in the UI) unless a site coordinate was
# published. Most local-opposition cancellations are NOT in the supply model
# (they were killed pre-construction), which is why the model-derived layer under-
# counts cancellations - these fill that gap. `match` = the operator's name in the
# model's Company column, so the page can show that operator's in-model GW context.
CITED_EVENTS = [
    # --- Indiana (a wave of local-opposition cancellations) ---
    {"company": "Google", "match": "Google",
     "project": "Franklin Township ('Deep Meadows')", "city": "Indianapolis", "state": "IN",
     "lat": 39.66, "lng": -86.00, "mw": None, "status": "cancelled",
     "orig_date": "2026-2028 (implied)", "new_date": None, "confidence": "high",
     "source": "Data Center Dynamics / IBJ", "pub_date": "2025-09-22",
     "url": "https://www.datacenterdynamics.com/en/news/google-withdraws-rezoning-proposal-for-468-acre-data-center-project-in-franklin-township-indianapolis/",
     "quote": "Google withdrew its rezoning petition for the $1bn, 468-acre Franklin Township site after community pushback, pre-empting a council vote it was set to lose."},
    {"company": "Undisclosed hyperscaler", "match": "",
     "project": "New Carlisle 'third' campus (rezoning denied)", "city": "New Carlisle", "state": "IN",
     "lat": 41.70, "lng": -86.51, "mw": None, "status": "cancelled",
     "orig_date": "2026-2028 (implied)", "new_date": None, "confidence": "high",
     "source": "WSBT / WNDU / WVPE", "pub_date": "2025-12-10",
     "url": "https://www.wndu.com/2025/12/10/st-joseph-county-council-meets-vote-data-center-proposal-new-carlisle/",
     "quote": "St. Joseph County Council denied the rezoning 7-2 for a ~$12-13bn, ~1,000-acre campus near the AWS New Carlisle site after ~10 hours of opposition."},
    {"company": "QTS", "match": "QTS",
     "project": "Porter County campus (withdrawn)", "city": "Valparaiso (Union Twp)", "state": "IN",
     "lat": 41.42, "lng": -87.18, "mw": None, "status": "cancelled",
     "orig_date": "2026-2028 (implied)", "new_date": None, "confidence": "high",
     "source": "Data Center Dynamics", "pub_date": "2025-09-01",
     "url": "https://www.datacenterdynamics.com/en/news/qts-cans-data-center-scheme-in-porter-county-indiana-after-protests/",
     "quote": "QTS dropped its $2bn, ~800-acre Porter County scheme: 'the conditions proposed would be economically challenging.'"},
    {"company": "Agincourt Investments", "match": "",
     "project": "Valparaiso project (option released)", "city": "Valparaiso", "state": "IN",
     "lat": 41.47, "lng": -87.06, "mw": None, "status": "cancelled",
     "orig_date": "2026-2028 (implied)", "new_date": None, "confidence": "high",
     "source": "Data Center Dynamics", "pub_date": "2025-03-26",
     "url": "https://www.datacenterdynamics.com/en/news/indianas-city-of-valparaiso-pulls-the-plug-on-agincourt-data-center-project/",
     "quote": "'Agincourt has agreed to withdraw from the project and to release its option on the land' (Mayor Jon Costas)."},
    {"company": "Surge Development", "match": "",
     "project": "Hancock County MegaSite (withdrawn)", "city": "Greenfield", "state": "IN",
     "lat": 39.79, "lng": -85.87, "mw": None, "status": "cancelled",
     "orig_date": "2026-2028 (implied)", "new_date": None, "confidence": "medium",
     "source": "WTHR / IBJ", "pub_date": "2025-01-01",
     "url": "https://www.wthr.com/article/money/business/application-withdrawn-surge-development-hancock-county-greenfield-data-center-tuttle-orchards/531-48204cbc-eba5-4666-bd66-6e88cb6d2412",
     "quote": "Surge withdrew its ~775-acre Hancock County rezoning after opposition; it is now eyeing Henry County instead."},
    # --- Other US cancellations ---
    {"company": "QTS / Compass", "match": "QTS",
     "project": "Prince William Digital Gateway (dead)", "city": "Gainesville/Manassas", "state": "VA",
     "lat": 38.83, "lng": -77.55, "mw": None, "status": "cancelled",
     "orig_date": "2026-2028 (implied)", "new_date": None, "confidence": "high",
     "source": "Virginia Business / WTOP / DCD", "pub_date": "2026-07-02",
     "url": "https://virginiabusiness.com/prince-william-digital-gateway-data-center-project-officially-dies/",
     "quote": "The ~2,100-acre, gigawatt-scale 'world's largest' campus 'is officially dead' after the last developer dropped its appeal and the zoning was ruled invalid."},
    {"company": "Microsoft", "match": "Microsoft",
     "project": "Caledonia site (rezoning withdrawn)", "city": "Caledonia (Racine Co.)", "state": "WI",
     "lat": 42.80, "lng": -87.89, "mw": None, "status": "cancelled",
     "orig_date": "2026-2028 (implied)", "new_date": None, "confidence": "medium",
     "source": "Racine County Eye / CNBC", "pub_date": "2025-10-08",
     "url": "https://racinecountyeye.com/2025/10/08/microsoft-abandon-1st-caledonia/",
     "quote": "Microsoft withdrew its Caledonia rezoning request; its separate $3.3bn Mount Pleasant campus continues (a later phase paused for redesign)."},
    {"company": "Tract", "match": "Tract",
     "project": "Mooresville Technology Park (withdrawn)", "city": "Mooresville (Iredell Co.)", "state": "NC",
     "lat": 35.58, "lng": -80.81, "mw": None, "status": "cancelled",
     "orig_date": "2026-2028 (implied)", "new_date": None, "confidence": "high",
     "source": "Data Center Dynamics / NBC News", "pub_date": "2025-09-15",
     "url": "https://www.datacenterdynamics.com/en/news/tract-pulls-plans-for-north-carolina-data-center-on-land-formerly-owned-by-nascar-legend-dale-earnhardt/",
     "quote": "Tract pulled its ~400-acre plan (on former Dale Earnhardt farmland) from the agenda after opposition, including from the Earnhardt family."},
    {"company": "Oracle / OpenAI (Stargate)", "match": "Oracle",
     "project": "Abilene 'Stargate' expansion (~600 MW)", "city": "Abilene (Taylor Co.)", "state": "TX",
     "lat": 32.45, "lng": -99.73, "mw": 600, "status": "cancelled",
     "orig_date": "2026-2028 (implied)", "new_date": None, "confidence": "medium",
     "source": "Bloomberg / Data Center Dynamics", "pub_date": "2026-03-06",
     "url": "https://www.bloomberg.com/news/articles/2026-03-06/oracle-and-openai-end-plans-to-expand-flagship-data-center",
     "quote": "Oracle and OpenAI ended plans for a ~600MW expansion of the Abilene flagship; the core ~1.2GW campus continues (Oracle disputes 'cancellation' framing)."},
    {"company": "Crusoe", "match": "Crusoe",
     "project": "Project Jade (Laramie Co., near Cheyenne)", "city": "Cheyenne", "state": "WY",
     "lat": 41.14, "lng": -104.82, "mw": 1800, "status": "cancelled",
     "orig_date": "2026-2028 (implied)", "new_date": None, "confidence": "high",
     "source": "Wyoming Tribune Eagle", "pub_date": "2026-06-09",
     "url": "https://www.wyomingnews.com/news/local_news/crusoe-pulls-out-of-project-jade-data-center/article_b9a8bde9-21c0-4e8a-9051-9154b91bf0eb.html",
     "quote": "Crusoe withdrew from the 1.8GW (scalable to 10GW) project; Tallgrass Energy is seeking a replacement tenant."},
    {"company": "Digital Realty", "match": "Digital Realty",
     "project": "2323 Bryan Street (downtown Dallas)", "city": "Dallas", "state": "TX",
     "lat": 32.79, "lng": -96.79, "mw": None, "status": "cancelled",
     "orig_date": "2026-2028 (implied)", "new_date": None, "confidence": "high",
     "source": "The Real Deal", "pub_date": "2025-06-10",
     "url": "https://therealdeal.com/texas/dallas/2025/06/10/digital-realty-cancels-104-million-dallas-data-center-project/",
     "quote": "Data-center capacity expansion cancelled; investment cut ~87% ($104M -> $13M) citing 'unforeseen circumstances and technical challenges'."},
    # --- Delays (originals 2026-2028 slipping) ---
    {"company": "STACK Infrastructure / Oracle (Stargate)", "match": "STACK Infrastructure",
     "project": "Project Jupiter (Dona Ana Co.)", "city": "Santa Teresa", "state": "NM",
     "lat": 31.87, "lng": -106.63, "mw": 1000, "status": "delayed",
     "orig_date": "1H2027", "new_date": "2029", "confidence": "high",
     "source": "SemiAnalysis", "pub_date": "2026-06-18",
     "url": "https://newsletter.semianalysis.com/p/stop-saying-half-of-2026-us-datacenter",
     "quote": "'Our base case moved first power off 2027 entirely and out to 2029' - gas-pipeline permitting, a state right-of-way denial and FERC review."},
    {"company": "Microsoft", "match": "Microsoft",
     "project": "Mount Pleasant (former Foxconn) - Phase 2", "city": "Mount Pleasant (Racine Co.)", "state": "WI",
     "lat": 42.72, "lng": -87.90, "mw": None, "status": "delayed",
     "orig_date": "2026-2027", "new_date": None, "confidence": "high",
     "source": "Data Center Dynamics", "pub_date": "2025-01-03",
     "url": "https://www.datacenterdynamics.com/en/news/microsoft-pauses-construction-on-part-of-data-center-site-in-mount-pleasant-wisconsin/",
     "quote": "'We have paused early construction work for this second phase while we evaluate scope and recent changes in technology.' Phase 1 continues."},
    {"company": "Microsoft", "match": "Microsoft",
     "project": "Licking County (New Albany / Heath / Hebron)", "city": "New Albany", "state": "OH",
     "lat": 40.08, "lng": -82.81, "mw": None, "status": "cancelled",
     "orig_date": "2025-2026", "new_date": None, "confidence": "high",
     "source": "Data Center Dynamics", "pub_date": "2025-04-08",
     "url": "https://www.datacenterdynamics.com/en/news/microsoft-backs-away-from-1bn-data-center-plans-in-licking-county-ohio/",
     "quote": "Microsoft paused ~$1bn across three Licking County sites; two of the three reverted to farmland use."},
    {"company": "Amazon / AWS", "match": "AWS",
     "project": "Becker (former Sherco site) - suspended", "city": "Becker (Sherburne Co.)", "state": "MN",
     "lat": 45.38, "lng": -93.88, "mw": None, "status": "delayed",
     "orig_date": "2026-2028 (implied)", "new_date": None, "confidence": "medium",
     "source": "Star Tribune / Data Center Dynamics", "pub_date": "2025-05-01",
     "url": "https://www.datacenterdynamics.com/en/news/amazon-web-services-suspends-plans-for-minnesota-data-center/",
     "quote": "Amazon suspended its Becker plan amid a fight over Minnesota's data-center tax exemption and a backup-generator permit ruling."},
]

# Portfolio / analyst-level pullbacks (aggregate, not tied to one site).
CITED_MACRO = [
    {"headline": "~25 US data-center projects cancelled in 2025 (vs 6 in 2024); Indiana ~8, ~5 GW",
     "detail": "Cancellations roughly quadrupled year-on-year on power/water limits and local opposition; Indiana tied Virginia for the most (~8 each), with a reported ~5 GW cancelled in Indiana alone.",
     "source": "Gizmodo / Heatmap (Baird)", "pub_date": "2025", "confidence": "medium",
     "url": "https://gizmodo.com/data-center-project-cancellations-quadrupled-in-2025-as-locals-fight-back-2000709669"},
    {"headline": "Microsoft: ~200 MW of US leases cancelled, up to ~2 GW deferred (US+Europe)",
     "detail": "TD Cowen: ~200MW of US colocation leases cancelled (Feb 2025) with at least two private operators, rising to up to 2GW deferred/cancelled plus >1GW of expired LOIs. Microsoft says it stays 'well positioned to meet increasing demand'.",
     "source": "Data Center Dynamics / TechCrunch (TD Cowen)", "pub_date": "2025-02-25", "confidence": "high",
     "url": "https://www.datacenterdynamics.com/en/news/microsoft-cancels-200mw-of-ai-data-center-leases-report/"},
    {"headline": "AWS: paused a portion of colocation leasing",
     "detail": "Wells Fargo: AWS paused some colocation leasing discussions ('particularly international ones'); signed leases were not cancelled. AWS calls it 'routine capacity management'.",
     "source": "CNBC / Data Center Dynamics (Wells Fargo)", "pub_date": "2025-04-21", "confidence": "high",
     "url": "https://www.cnbc.com/2025/04/21/amazon-has-paused-some-data-center-lease-commitments-wells-fargo.html"},
    {"headline": "$130B / 75+ US buildouts blocked or delayed in Q1 2026",
     "detail": "Data Center Watch: more than 75 projects worth ~$130bn blocked or delayed in the first four months of 2026 amid bipartisan local opposition over power and water costs.",
     "source": "Tom's Hardware / Fortune (Data Center Watch)", "pub_date": "2026", "confidence": "medium",
     "url": "https://www.tomshardware.com/tech-industry/artificial-intelligence/more-than-75-data-center-build-outs-worth-usd130-billion-have-been-successfully-blocked-in-the-first-four-months-of-2026-bipartisan-opposition-mounts-nationwide-over-fears-of-soaring-power-and-water-costs"},
    {"headline": "Caveat: 'more than half of 2026 US DCs cancelled/delayed' is disputed",
     "detail": "SemiAnalysis argues the 'half of 2026 capacity is cancelled' framing is a denominator error - its own year-end 2026 NA self-build forecast moved only ~1% (colo <5%). Much of the slip is delay (e.g. an electrical-gear shortage), not demand collapse.",
     "source": "SemiAnalysis", "pub_date": "2026-06-18", "confidence": "medium",
     "url": "https://newsletter.semianalysis.com/p/stop-saying-half-of-2026-us-datacenter"},
]


DELAY_MIN_MONTHS = 3.0        # a slip must exceed this to count as "delayed"
DELAY_GRADIENT_CAP = 48.0     # months mapped to the reddest end of the yellow->red scale


def extract_delayed_cancelled(wb, utilization, pue_2028):
    """US facilities whose ORIGINAL online date (col BY, 'Start of operations')
    fell in 2026-2028, that are now cancelled or delayed in the model.

    Scope is exactly the 2026-2028 original cohort (the user's ask). Within it:
      * cancelled - model Live Date (col CC) is a far-future sentinel (>=2034) or
        blank while capacity is still planned  -> plotted RED on the map.
      * delayed   - CC lands more than DELAY_MIN_MONTHS after BY -> plotted on a
        yellow->red gradient scaled by the slip in months.
    Each facility carries lat/lng (cols CN/CO) so the page can draw them on the
    same offline US map the geography page uses. Energy uses the SAME conversion
    as derive_energy_split (8760h x utilization x PUE) so TWh reconciles with
    pages 3 & 7. Externally-reported, cited events (CITED_EVENTS) are merged in
    as a separate labelled layer - including sites the model does not carry.
    """
    import datetime
    ws = wb["NA Data Center Supply"]
    cA, cCountry = col("A"), col("CT")
    cCE, cCD, cBV, cBY, cCC = col("CE"), col("CD"), col("BV"), col("BY"), col("CC")
    cLat, cLng = col("CN"), col("CO")
    cCompany, cState, cCity, cType = col("CV"), col("CP"), col("CQ"), col("CX")
    cGpu, cTenant = col("DA"), col("DB")               # for the deployment-class heuristic
    maxcol = max(cA, cCountry, cCE, cCD, cBV, cBY, cCC, cLat, cLng,
                 cCompany, cState, cCity, cType, cGpu, cTenant)

    intensity = round(HOURS_PER_YEAR * utilization * pue_2028 / 1000.0, 3)  # TWh per GW-yr

    facilities = []
    tot = {"delayed": 0.0, "cancelled": 0.0}
    n = {"delayed": 0, "cancelled": 0}
    # capacity split into committed (Under Construction, col BV) vs paper (Planned, col CD);
    # CE == BV + CD in this sheet, so uc + planned reconciles to the class total.
    by_phase = {"delayed": {"uc": 0.0, "planned": 0.0},
                "cancelled": {"uc": 0.0, "planned": 0.0}}
    # For the shortfall test: does the slip actually remove capacity the client's
    # waterfall counts on? Split UC vs Planned by whether the REVISED live date
    # still lands in-window (<=2028), slips past 2028, or is cancelled.
    sf_buckets = ("by2028", "past2028", "cancelled", "unknown")
    sf = {"uc": {k: 0.0 for k in sf_buckets},
          "planned": {k: 0.0 for k in sf_buckets}}
    by_company = {}
    buckets = {"3-6 mo": 0, "6-12 mo": 0, "1-2 yr": 0, "2-4 yr": 0, ">4 yr / cancelled": 0}
    delays = []
    n_no_coord = 0

    def yof(v):
        return v.year if isinstance(v, datetime.datetime) else None

    for row in ws.iter_rows(min_row=6, values_only=True):
        if len(row) <= maxcol:
            continue
        a = row[cA]
        if not (isinstance(a, str) and len(a) >= 8 and "-" in a):
            continue
        country = row[cCountry]
        if not (isinstance(country, str) and country.strip().upper() in ("USA", "US", "UNITED STATES")):
            continue
        by = row[cBY]
        if yof(by) not in (2026, 2027, 2028):              # scope: ORIGINAL online 2026-2028
            continue
        gw = num(row[cCE]) / 1000.0                        # Planned+UC, MW->GW
        uc_gw = num(row[cBV]) / 1000.0                      # Under Construction (col BV)
        planned_gw = num(row[cCD]) / 1000.0                 # Planned (col CD); CE == BV + CD
        cc = row[cCC]
        ccyr = yof(cc)

        # Reconcile the complete original-2026-28 cohort before the map-specific
        # delay and coordinate filters below.  This makes shortfall_split a true
        # time-aligned capacity bridge rather than a subset of delayed sites.
        if (ccyr is not None and ccyr >= 2034) or (ccyr is None and planned_gw > 0):
            sf_bucket = "cancelled"
        elif ccyr is not None:
            sf_bucket = "past2028" if ccyr >= 2029 else "by2028"
        else:
            sf_bucket = "unknown"
        sf["uc"][sf_bucket] += uc_gw
        sf["planned"][sf_bucket] += planned_gw

        # classify the affected facilities used by the delayed/cancelled pages
        if sf_bucket == "cancelled":
            status, delay_mo = "cancelled", None
        elif isinstance(cc, datetime.datetime) and isinstance(by, datetime.datetime):
            dm = (cc - by).days / 30.44
            if dm <= DELAY_MIN_MONTHS:
                continue                                   # on-time / trivial slip -> not shown
            status, delay_mo = "delayed", round(dm, 1)
        else:
            continue
        if gw <= 0 and status == "delayed":
            continue

        lat, lng = row[cLat], row[cLng]
        if not (isinstance(lat, (int, float)) and isinstance(lng, (int, float))):
            n_no_coord += 1
            continue

        co = row[cCompany].strip() if isinstance(row[cCompany], str) else "Unknown"
        st = row[cState].strip() if isinstance(row[cState], str) else ""
        city = row[cCity].strip() if isinstance(row[cCity], str) else ""
        tot[status] += gw
        n[status] += 1
        by_phase[status]["uc"] += uc_gw
        by_phase[status]["planned"] += planned_gw
        by_company[co] = by_company.get(co, 0.0) + gw
        if status == "cancelled":
            buckets[">4 yr / cancelled"] += 1
        else:
            delays.append(delay_mo)
            if delay_mo < 6:
                buckets["3-6 mo"] += 1
            elif delay_mo < 12:
                buckets["6-12 mo"] += 1
            elif delay_mo < 24:
                buckets["1-2 yr"] += 1
            elif delay_mo < 48:
                buckets["2-4 yr"] += 1
            else:
                buckets[">4 yr / cancelled"] += 1
        facilities.append({
            "lat": round(float(lat), 4), "lng": round(float(lng), 4),
            "co": co, "city": city, "st": st,
            "gw": round(gw, 4), "orig": yof(by), "rev": ccyr,
            "delay_mo": delay_mo, "status": status,
            "uc": uc_gw > 0,                                # committed (Under Construction) vs Planned
            "cls": _deploy_class(row[cCompany], row[cTenant], row[cGpu], row[cType]),
        })

    def r3(v):
        return round(v, 3)

    facilities.sort(key=lambda f: -f["gw"])                # small circles drawn on top
    delays.sort()
    ndel = len(delays)

    def top(d, k=12):
        return [{"name": kk, "gw": r3(vv)} for kk, vv in sorted(d.items(), key=lambda x: -x[1])[:k]]

    # cited events: keep only those in scope (original 2026-2028) for the map layer;
    # attach the model's operator-level delayed+cancelled GW as company-level context.
    cited = []
    for ev in CITED_EVENTS:
        e = dict(ev)
        e["model_company_gw"] = r3(by_company.get(ev.get("match", ""), 0.0))
        cited.append(e)

    return {
        "scope": "US facilities with an original online date (Start of operations) in 2026-2028",
        "facilities": facilities,
        "n_facilities": len(facilities),
        "n_no_coord": n_no_coord,
        "totals_gw": {"delayed": r3(tot["delayed"]), "cancelled": r3(tot["cancelled"]),
                      "total": r3(tot["delayed"] + tot["cancelled"])},
        "by_phase": {k: {"uc": r3(v["uc"]), "planned": r3(v["planned"])} for k, v in by_phase.items()},
        "uc_total_gw": r3(by_phase["delayed"]["uc"] + by_phase["cancelled"]["uc"]),
        "planned_total_gw": r3(by_phase["delayed"]["planned"] + by_phase["cancelled"]["planned"]),
        "shortfall_split": {ph: {k: r3(v) for k, v in d.items()} for ph, d in sf.items()},
        "n_sites": n,
        "delay_stats": {"median_mo": delays[ndel // 2] if ndel else None,
                        "max_mo": delays[-1] if ndel else None,
                        "buckets": buckets},
        "delay_gradient_cap_mo": DELAY_GRADIENT_CAP,
        "intensity_twh_per_gw": intensity,
        "utilization": utilization, "pue_2028": pue_2028, "hours": HOURS_PER_YEAR,
        "twh": {"delayed_deferred": r3(tot["delayed"] * intensity),
                "cancelled_lost": r3(tot["cancelled"] * intensity),
                "total": r3((tot["delayed"] + tot["cancelled"]) * intensity)},
        "by_company": top(by_company),
        "cited_events": cited,
        "cited_macro": CITED_MACRO,
    }


# --------------------------------------------------------------------------- #
# Derivations: energy split (TWh by chip class) and compute split (GPU vs ASIC)
# --------------------------------------------------------------------------- #
# Watts per volume row (Power Summary row number -> W per chip).
# 'sourced' = taken from the model file's 'Power Requirements per Chip' /
# MS 'GPU Roadmap'; others are labeled analogs (assumptions).
ENERGY_WATTS = {
    # GPU rows
    24: ("H100", 700, True), 25: ("A100", 400, True), 26: ("B100/B200", 1000, True),
    27: ("R100 Rubin", 2300, True), 28: ("Rubin Ultra", 3600, True), 29: ("Feynman", 7000, True),
    30: ("R100+3", 3600, False), 31: ("Other NVDA", 500, False), 32: ("AMD Other", 500, False),
    # AMD MI-series: named rows found in 'Power Requirements per Chip' rows 99-105
    # (a prior version of this table guessed these; corrected against the sourced values)
    33: ("MI300X", 750, True), 34: ("MI325X", 1000, True), 35: ("MI350", 1400, True),
    36: ("MI400", 2500, True), 37: ("MI500 (inferred '+1' gen)", 2600, True), 38: ("MI '+2' gen (extrapolated)", 2700, False),
    # Intel: named rows 82-86 (Gaudi guesses happened to match; Falcon Shores corrected)
    39: ("Habana Gaudi2", 600, True), 40: ("Habana Gaudi3", 900, True), 41: ("Falcon Shores", 1200, True),
    42: ("Future Intel", 1000, False),
    # ASIC rows (r57 China/Ascend excluded upstream)
    45: ("TPU pre-v5 (v4 Pufferfish)", 200, True), 46: ("TPU v5p Viperfish", 550, True), 47: ("TPU v5e Viperlite", 300, True),
    48: ("TPU v6 Ghostlite", 390, True), 49: ("TPU v7 Ironwood", 980, True), 50: ("TPU v8AX Sunfish", 1200, True),
    51: ("TPU v9 Pumafish", 1500, True),
    # AWS: rows 75-81, 87-92 give named Trainium/Inferentia gens once the full
    # sheet (not just rows 11-50) is scanned
    52: ("Trainium2 Teton PD", 500, True), 53: ("Inferentia2/Trainium1", 200, True),
    54: ("Trainium3 Teton PD", 570, True), 55: ("Trainium4 (UALink/NVLink)", 1200, True),
    56: ("Trainium '2+3' gen (extrapolated)", 1400, False),
    # ByteDance/Meta/OpenAI: named rows 54-63 give real (and surprising) values —
    # ByteDance and early MTIA are far lower-power than the flagship-chip guess
    # this table used before; OpenAI's own 'Titan' chip is much higher-power
    58: ("ByteDance Custom ASIC Gen1/2", 90, True),
    59: ("Meta MTIA (Gen3-5 avg)", 1000, True),
    60: ("OpenAI 'Titan' (avg Titan1/2)", 1800, True),
    61: ("Microsoft Maia (avg of 5 named gens)", 962, True),
    62: ("xAI", 800, False),  # no named xAI chip found anywhere in the file
}
HOURS_PER_YEAR = 8760.0


def _watt_base(vol_rows, cumulative):
    """Sum of units x W per year (GW of chip power). cumulative=True uses the trailing
    install base (chips keep running once bought; no retirement within the window)."""
    n = 6                                                   # 2023..2028
    base = [0.0] * n
    for vr in vol_rows:
        w = ENERGY_WATTS.get(vr["row"])
        if w is None:
            continue
        run = 0.0
        for i in range(n):
            units_k = vr["values"][i]
            if cumulative:
                run += units_k
                base[i] += run * 1000 * w[1] / 1e9          # GW
            else:
                base[i] += units_k * 1000 * w[1] / 1e9
    return base


def derive_energy_split(client_chips):
    """TWh by chip class = install base x W x utilization x 8760h x PUE.
    Cumulative install base (no retirement). Validated against the client's
    own aggregate GenAI TWh (Power Summary r10)."""
    util = client_chips["utilization"]
    pue = client_chips["pue"]
    gpu_gw = _watt_base(client_chips["gpu_vol_rows"], cumulative=True)
    asic_gw = _watt_base(client_chips["asic_vol_rows"], cumulative=True)
    to_twh = lambda gw_list: [round(g * util * p * HOURS_PER_YEAR / 1000.0, 1)
                              for g, p in zip(gw_list, pue)]
    gpu_twh, asic_twh = to_twh(gpu_gw), to_twh(asic_gw)
    derived_total = [round(a + b, 1) for a, b in zip(gpu_twh, asic_twh)]
    client_total = client_chips["twh_genai_global"]
    ratio = [round(d / c, 2) if c else None for d, c in zip(derived_total, client_total)]
    return {"years": client_chips["years"], "gpu_twh": gpu_twh, "asic_twh": asic_twh,
            "derived_total_twh": derived_total, "client_total_twh": client_total,
            "validation_ratio": ratio, "utilization": util, "pue": pue}


def derive_compute_split(client_chips):
    """ASIC ExaFLOPs estimated at efficiency parity with same-year Nvidia:
    asic_exa[y] = nv_exa[y] x (asic annual watt-base / nvidia annual watt-base).
    Nvidia watt-base uses the Nvidia-only volume rows (24-31)."""
    nv_rows = [vr for vr in client_chips["gpu_vol_rows"] if vr["row"] <= 31]
    nv_gw = _watt_base(nv_rows, cumulative=False)
    asic_gw = _watt_base(client_chips["asic_vol_rows"], cumulative=False)
    # exa years are 2024-2028 -> volume years index 1..5
    nv_exa = client_chips["exa_total"]
    asic_exa = []
    for i, exa in enumerate(nv_exa):
        vi = i + 1                                          # 2024 -> volume index 1
        asic_exa.append(round(exa * asic_gw[vi] / nv_gw[vi], 1) if nv_gw[vi] else 0.0)
    return {"years": client_chips["exa_gen_years"], "nv_exa": nv_exa, "asic_exa_parity": asic_exa}


# Generation-level pairing of the two independent watts-per-chip tables.
# (client_substring, ms_substring) matched case-insensitively; None = absent.
WATTS_PAIRS = [
    ("A100", "a100", "a100"),
    ("H100 (SXM)", "h100 sxm", "h100"),
    ("B200", "b200 ", "b200"),
    ("GB200 NVL72", "gb200 nvl72", "gb200 nvl72"),
    ("GB300 NVL72", "gb300 nvl72", "gb300 nvl72"),
    ("VR200 NVL144", "vr200 nvl144", "vr200 nvl144"),
    ("VR300 / R300", "r300", "vr300"),
    ("F200", "f200", None),
]


def build_watts_compare(model_watts, ms_tdp):
    def find(items, key_field, sub):
        if sub is None:
            return None
        for it in items:
            s = it[key_field].lower()
            # exact-ish: substring match but avoid 'CPX' variants when not asked for
            if sub in s and ("cpx" in sub or "cpx" not in s):
                return it["watts"]
        return None

    out = []
    for gen, c_sub, m_sub in WATTS_PAIRS:
        cw = find(model_watts, "sku", c_sub)
        mw = find(ms_tdp, "sku", m_sub)
        if cw is not None or mw is not None:
            out.append({"gen": gen, "client": cw, "ms": mw})
    return out


BRIDGE_PUE = 1.3


def derive_bridge(nv, client_chips, supply):
    """Chip-implied DC power: racks/yr x per-rack usage kW x PUE -> GW/yr (global racks).

    VR300 has no rack-power figure in the NV file; approximated at VR200 kW.
    HGX/DGX 8-GPU servers are excluded (rack tracker covers NVL72-scale only),
    so this understates total chip-driven demand.
    """
    kw = dict(nv["rack_kw"])
    kw["VR300"] = kw["VR200 NVL72"]
    years = nv["rack_years"]                                   # [2025, 2026, 2027]
    implied = []
    for i in range(len(years)):
        mw = sum(s["values"][i] * kw[s["name"]] for s in nv["rack_by_sku"]) / 1000.0
        implied.append(round(mw * BRIDGE_PUE / 1000.0, 2))     # GW incl. PUE

    cy = client_chips["years"]
    client_adds = [client_chips["us_additions_gw"][cy.index(y)] for y in years]
    st = supply["year_series_gw"]["total"]
    sy = supply["years"]
    model_adds = [round(st[sy.index(y)] - st[sy.index(y - 1)], 2) for y in years]

    return {"years": years, "pue": BRIDGE_PUE, "rack_kw": kw,
            "implied_gw": implied, "client_us_additions_gw": client_adds,
            "model_us_additions_gw": model_adds}


# --------------------------------------------------------------------------- #
# Reconciliation
# --------------------------------------------------------------------------- #
def reconcile(shortfall, supply, delayed_cancelled=None, balance=None):
    needed = shortfall["shortfall_before_solutions"]
    us_uc = supply["totals_gw"]["uc"]
    us_pluc = supply["totals_gw"]["pluc"]
    us_planned = supply["totals_gw"]["planned"]
    us_add_26_28 = supply["totals_gw"]["delta_25_28"]

    out = {
        "needed_shortfall": needed,                       # 37.71
        "client_uc": abs(shortfall["less_under_construction"]),   # 14.85
        "model_us_uc": us_uc,                             # ~26.28
        "client_demand": shortfall["demand_2026_28"],     # 67.56
        "model_us_additions_26_28": us_add_26_28,         # ~94.11
        "model_us_planned": us_planned,                   # ~202
        "model_us_pluc": us_pluc,                         # ~228
        "residual_pluc_vs_needed": round(us_pluc - needed, 2),
        # Descriptive pipeline multiple only.  This is intentionally not called
        # coverage because Planned+UC is an all-year stock while needed is a
        # 2026-28 residual after UC and grid access have already been credited.
        "pipeline_to_shortfall_multiple": round(us_pluc / needed, 2) if needed else None,
        "coverage_ratio_pluc": round(us_pluc / needed, 2) if needed else None,
        "coverage_ratio_additions": round(us_add_26_28 / shortfall["demand_2026_28"], 2)
        if shortfall["demand_2026_28"] else None,
    }

    if delayed_cancelled is not None:
        sf = delayed_cancelled["shortfall_split"]
        uc_by2028 = sf["uc"]["by2028"]
        planned_by2028 = sf["planned"]["by2028"]
        by2028_total = uc_by2028 + planned_by2028
        incremental = max(0.0, uc_by2028 - out["client_uc"]) + planned_by2028
        residual = max(0.0, needed - incremental)
        slips = sf["uc"]["past2028"] + sf["planned"]["past2028"]
        cancelled = sf["uc"]["cancelled"] + sf["planned"]["cancelled"]
        out.update({
            "model_us_by2028_uc_gw": round(uc_by2028, 3),
            "model_us_by2028_planned_gw": round(planned_by2028, 3),
            "model_us_by2028_total_gw": round(by2028_total, 3),
            "model_us_incremental_vs_client_gw": round(incremental, 3),
            "time_aligned_residual_gw": round(residual, 3),
            "time_aligned_coverage_ratio": round(incremental / needed, 2) if needed else None,
            "model_us_slips_past_2028_gw": round(slips, 3),
            "model_us_cancelled_gw": round(cancelled, 3),
        })

    if balance is not None and 2028 in balance["years"]:
        i28 = balance["years"].index(2028)
        demand_2028 = balance["total_demand_mw"][i28] / 1000.0
        capacity_2028 = balance["total_capacity_mw"][i28] / 1000.0
        out.update({
            "model_na_annual_balance_2028_gw": round(balance["surplus_deficit_mw"][i28] / 1000.0, 3),
            "model_na_cum_balance_2028_gw": round(balance["cum_surplus_deficit_mw"][i28] / 1000.0, 3),
            "model_na_demand_2028_gw": round(demand_2028, 3),
            "model_na_capacity_2028_gw": round(capacity_2028, 3),
            "model_na_demand_minus_capacity_2028_gw": round(demand_2028 - capacity_2028, 3),
        })

    return out


# --------------------------------------------------------------------------- #
# Self-check (fails loudly if a cell reference drifts)
# --------------------------------------------------------------------------- #
def self_check(shortfall, supply, nv=None, buildout=None, client_chips=None, extra_checks=None):
    checks = [
        ("shortfall 37.71", shortfall["shortfall_before_solutions"], 37.71, 0.1),
        ("demand 67.56", shortfall["demand_2026_28"], 67.56, 0.1),
        ("US UC 26.28", supply["totals_gw"]["uc"], 26.28, 0.5),
        ("US Planned+UC 228.38", supply["totals_gw"]["pluc"], 228.38, 1.0),
        ("US additions 26-28 94.11", supply["totals_gw"]["delta_25_28"], 94.11, 1.0),
    ]
    if nv is not None:
        gb300_2026 = next(s for s in nv["rack_by_sku"] if s["name"] == "GB300 NVL72")["values"][1]
        checks += [
            ("NV racks total 2025 = 28918", nv["rack_totals"][0], 28918, 2),
            ("NV GB300 racks 2026 = 73260", gb300_2026, 73260, 2),
            ("NV rack kW GB200 = 131.67", nv["rack_kw"]["GB200 NVL72"], 131.67, 0.1),
        ]
    if buildout is not None:
        msft = next(h for h in buildout["hyperscalers"] if h["name"] == "Microsoft")
        sb_na_2025 = msft["sub"]["Self-build - North America"][HN_YEARS.index(2025)]
        checks.append(("MSFT Self-build NA 2025 = 4485", sb_na_2025, 4485.1, 5))
    if client_chips is not None:
        checks += [
            ("Total GPUs 2028 = 9701k", client_chips["total_gpus"][-1], 9701.05, 2),
            ("US DC power 2028 = 119.4 GW", client_chips["us_power_gw"][-1], 119.4, 0.5),
        ]
    if extra_checks:
        checks += extra_checks
    problems = []
    for name, got, want, tol in checks:
        if abs(got - want) > tol:
            problems.append(f"  {name}: expected ~{want}, got {got}")
    if problems:
        raise SystemExit("SELF-CHECK FAILED (cell references may have drifted):\n" + "\n".join(problems))
    print("Self-check passed: all anchor values match planning figures.")


# --------------------------------------------------------------------------- #
# HTML rendering
# --------------------------------------------------------------------------- #
def render_html(data, template_file):
    with open(VENDOR_CHARTJS, "r", encoding="utf-8") as f:
        chartjs = f.read()
    with open(template_file, "r", encoding="utf-8") as f:
        template = f.read()
    payload = json.dumps(data)
    html = template.replace("/*__CHARTJS__*/", chartjs).replace("__DATA__", payload)

    def _read(p):
        with open(p, "r", encoding="utf-8") as fh:
            return fh.read()

    # D3 assets (geo + flows pages). Each placeholder is inlined only if present.
    # d3.min.js must precede d3-sankey (sankey reads d3-array/d3-shape off the global).
    if "/*__D3__*/" in html:
        html = html.replace("/*__D3__*/", _read(VENDOR_D3))
    if "/*__SANKEY__*/" in html:
        html = html.replace("/*__SANKEY__*/", _read(VENDOR_SANKEY))
    # Map data (geo page only).
    if "/*__TOPOJSON__*/" in html:
        html = html.replace("/*__TOPOJSON__*/", _read(VENDOR_TOPOJSON))
    if "__USTOPO__" in html:
        html = html.replace("__USTOPO__", _read(VENDOR_US_TOPO))
    if "__PIPES__" in html:
        html = html.replace("__PIPES__", _read(VENDOR_PIPELINES))
    return html


def main():
    for p in (CLIENT_FILE, MODEL_FILE, NV_FILE, VENDOR_CHARTJS, TEMPLATE_FILE,
              CHIPS_TEMPLATE_FILE, COMPUTE_TEMPLATE_FILE, SYNTHESIS_TEMPLATE_FILE,
              GEO_TEMPLATE_FILE, FLOWS_TEMPLATE_FILE, CHIPPOWER_TEMPLATE_FILE,
              DELAYED_TEMPLATE_FILE, CANCELLEDPROFILE_TEMPLATE_FILE, SHORTFALLMATH_TEMPLATE_FILE,
              VENDOR_D3, VENDOR_SANKEY, VENDOR_TOPOJSON, VENDOR_US_TOPO, VENDOR_PIPELINES):
        if not os.path.exists(p):
            raise SystemExit(f"Missing required file: {p}")

    print("Extracting client file (shortfall + chips) ...")
    client_wb = openpyxl.load_workbook(CLIENT_FILE, read_only=True, data_only=True)
    shortfall = extract_shortfall(client_wb)
    client_chips = extract_client_chips(client_wb)
    ai_power_us = extract_ai_power_us(client_wb)
    chip_power_global = extract_chip_power_global(client_wb)
    client_wb.close()

    print("Extracting model file (supply + buildout + chip watts) ...")
    model_wb = openpyxl.load_workbook(MODEL_FILE, read_only=True, data_only=True)
    supply = extract_supply(model_wb)
    schedule_bridge = extract_schedule_bridge(model_wb)
    buildout = extract_buildout(model_wb)
    model_watts = extract_chip_watts(model_wb)
    geo = extract_geo(model_wb)
    cap_by_deploy = extract_capacity_by_deploy(model_wb)
    print("Computing natural-gas pipeline proximity ...")
    pipes = load_pipelines()
    annotate_pipe_proximity(geo, pipes)
    print("Extracting model file (demand, balance, revisions, slippage, additions) ...")
    customer = extract_customer_demand(model_wb)
    unconstrained = extract_unconstrained(model_wb)
    balance = extract_na_balance(model_wb)
    revisions = extract_revisions(model_wb)
    slippage = extract_slippage(model_wb)
    dash_adds = extract_dashboard_adds(model_wb)
    delayed_cancelled = extract_delayed_cancelled(
        model_wb, client_chips["utilization"], client_chips["pue"][-1])
    # pipeline proximity for the cancelled/delayed facilities (same network as geo page)
    annotate_pipe_proximity({"facilities": delayed_cancelled["facilities"]}, pipes)
    model_wb.close()

    print("Extracting NV server model (chips, racks, customers) ...")
    nv = extract_nv(NV_FILE)

    extra = [
        ("constrained AI power 2028 = 110410 MW", customer["ai_power_mw"][customer["years"].index(2028)], 110409.6, 5),
        ("unconstrained AI power 2028 = 126407 MW", unconstrained["ai_power_mw"][unconstrained["years"].index(2028)], 126406.8, 5),
        ("NA balance surplus 2028 = +7711 MW", balance["surplus_deficit_mw"][balance["years"].index(2028)], 7711.0, 5),
        ("NA cumulative balance 2028 = -689 MW", balance["cum_surplus_deficit_mw"][balance["years"].index(2028)], -689.0, 5),
        ("NA total demand 2028 = 89913 MW", balance["total_demand_mw"][balance["years"].index(2028)], 89912.8, 5),
        ("GenAI TWh 2028 = 801.64", client_chips["twh_genai_global"][-1], 801.64, 0.5),
        ("US GenAI TWh 2028 = 455.07", client_chips["twh_genai_us"][-1], 455.07, 0.5),
        ("training share 2028 = 1.3228", client_chips["training_share"][-1], 1.3228, 0.01),
        ("ExaFLOPs total 2028 = 570733", client_chips["exa_total"][-1], 570733.1, 2),
        ("slippage median = 9.0 mo", slippage["median_mo"], 9.0, 0.5),
        ("vintage current 2026 = 100913", revisions["vintages"][-1]["y2026"], 100913, 2),
        ("vintage first 2030 = 140851", revisions["vintages"][0]["y2030"], 140851, 2),
        ("Oracle pct 2030 = -0.435", revisions["oracle_pct"][-1], -0.435, 0.005),
        ("AI Power_US DC cap 2028 = 69.26 GW", ai_power_us["dc_power_gw"][-1], 69.264, 0.5),
        ("cap-by-deploy 2028 == supply year-end total",
         sum(cap_by_deploy["annual"][k][cap_by_deploy["years"].index(2028)]
             for k in ("selfbuilt", "leasing", "neocloud")),
         supply["year_series_gw"]["total"][supply["years"].index(2028)], 0.5),
        ("Global chip DC power 2028 = 122 GW", chip_power_global["dc_power_gw"][-1], 122.0, 1.0),
        ("Global chip TWh actual 2028 = 801.6", chip_power_global["twh_actual"][-1], 801.6, 1.0),
        ("Global chip server power 2028 ~= 25.5 GW", chip_power_global["total_servers_gw"][-1], 25.5, 2.0),
        ("delayed (orig 2026-28) GW ~= 139.9", delayed_cancelled["totals_gw"]["delayed"], 139.9, 3.0),
        ("cancelled (orig 2026-28) GW ~= 27.2", delayed_cancelled["totals_gw"]["cancelled"], 27.2, 3.0),
        ("mapped facilities (orig 2026-28) ~= 855",
         delayed_cancelled["n_facilities"], 855, 15),
        ("cancelled + delayed-past-2028 GW ~= 119.8",
         delayed_cancelled["totals_gw"]["cancelled"]
         + delayed_cancelled["shortfall_split"]["uc"]["past2028"]
         + delayed_cancelled["shortfall_split"]["planned"]["past2028"], 119.8, 3.0),
        ("original-2026-28 cohort landing by 2028 ~= 47.4 GW",
         delayed_cancelled["shortfall_split"]["uc"]["by2028"]
         + delayed_cancelled["shortfall_split"]["planned"]["by2028"], 47.4, 1.0),
        ("annual-schedule bridge starts at 94.1 GW",
         schedule_bridge["annual_total_gw"], 94.113, 0.01),
        ("annual-schedule bridge lands at 47.4 GW",
         schedule_bridge["revised_cohort_lands_gw"], 47.387, 0.01),
        ("annual-schedule bridge arithmetic reconciles",
         schedule_bridge["annual_total_gw"]
         - schedule_bridge["annual_by_revised_gw"]["slips_past_2028"]
         - schedule_bridge["annual_by_revised_gw"]["cancelled"]
         - schedule_bridge["annual_by_revised_gw"]["no_revised_date"]
         - schedule_bridge["outside_original_cohort_gw"]
         + schedule_bridge["phase_capacity_adjustment_gw"],
         schedule_bridge["revised_cohort_lands_gw"], 0.01),
    ]
    self_check(shortfall, supply, nv, buildout, client_chips, extra)

    rec = reconcile(shortfall, supply, delayed_cancelled, balance)
    rec["schedule_bridge"] = schedule_bridge
    bridge = derive_bridge(nv, client_chips, supply)
    energy_split = derive_energy_split(client_chips)
    compute_split = derive_compute_split(client_chips)
    chips_by_deploy = derive_chips_by_deploy(ai_power_us, buildout)
    data = {"shortfall": shortfall, "supply": supply, "reconciliation": rec,
            "chips": {"nv": nv, "client": client_chips, "model_watts": model_watts,
                      "watts_compare": build_watts_compare(model_watts, nv["ms_tdp"]),
                      "buildout": buildout, "bridge": bridge},
            "compute": {"energy_split": energy_split, "compute_split": compute_split,
                        "customer": customer},
            "synthesis": {"unconstrained": unconstrained, "balance": balance,
                          "revisions": revisions, "slippage": slippage,
                          "dash_adds": dash_adds},
            "geo": {"facilities": geo, "capacity_by_deploy": cap_by_deploy,
                    "chips_by_deploy": chips_by_deploy, "ai_power_us": ai_power_us},
            "chippower": {"global": chip_power_global, "us": ai_power_us},
            "delayed_cancelled": delayed_cancelled,
            "meta": {
                "client_source": "Morgan Stanley",
                "model_source": "SemiAnalysis",
                "client_file": os.path.basename(CLIENT_FILE),
                "model_file": os.path.basename(MODEL_FILE),
                "nv_file": os.path.basename(NV_FILE),
                "client_asof": "June 11, 2026",
                "model_asof": "April 7",
            }}

    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

    with open(OUT_HTML, "w", encoding="utf-8") as f:
        f.write(render_html(data, TEMPLATE_FILE))

    with open(OUT_CHIPS_HTML, "w", encoding="utf-8") as f:
        f.write(render_html(data, CHIPS_TEMPLATE_FILE))

    with open(OUT_COMPUTE_HTML, "w", encoding="utf-8") as f:
        f.write(render_html(data, COMPUTE_TEMPLATE_FILE))

    with open(OUT_SYNTHESIS_HTML, "w", encoding="utf-8") as f:
        f.write(render_html(data, SYNTHESIS_TEMPLATE_FILE))

    with open(OUT_GEO_HTML, "w", encoding="utf-8") as f:
        f.write(render_html(data, GEO_TEMPLATE_FILE))

    with open(OUT_FLOWS_HTML, "w", encoding="utf-8") as f:
        f.write(render_html(data, FLOWS_TEMPLATE_FILE))

    with open(OUT_CHIPPOWER_HTML, "w", encoding="utf-8") as f:
        f.write(render_html(data, CHIPPOWER_TEMPLATE_FILE))

    with open(OUT_DELAYED_HTML, "w", encoding="utf-8") as f:
        f.write(render_html(data, DELAYED_TEMPLATE_FILE))

    with open(OUT_CANCELLEDPROFILE_HTML, "w", encoding="utf-8") as f:
        f.write(render_html(data, CANCELLEDPROFILE_TEMPLATE_FILE))

    with open(OUT_SHORTFALLMATH_HTML, "w", encoding="utf-8") as f:
        f.write(render_html(data, SHORTFALLMATH_TEMPLATE_FILE))

    print("\n=== Reconciliation summary (GW) ===")
    print(f"  Needed (Power Shortfall before solutions): {rec['needed_shortfall']}")
    print(f"  Client Under-Construction assumption:      {rec['client_uc']}")
    print(f"  Model US Under-Construction:               {rec['model_us_uc']}")
    print(f"  Client demand 2026-28:                     {rec['client_demand']}")
    print(f"  Model US capacity added 2026-28:           {rec['model_us_additions_26_28']}")
    print(f"  Model US Planned:                          {rec['model_us_planned']}")
    print(f"  Model US Planned + UC:                     {rec['model_us_pluc']}")
    print(f"  Planned+UC vs needed (all-year exposure):  {rec['residual_pluc_vs_needed']}  "
          f"(pipeline multiple x{rec['pipeline_to_shortfall_multiple']})")
    print(f"  Status-adjusted cohort landing by 2028:    {rec['model_us_by2028_total_gw']}")
    print(f"  Incremental vs client-credited UC:         {rec['model_us_incremental_vs_client_gw']}  "
          f"(x{rec['time_aligned_coverage_ratio']}, residual {rec['time_aligned_residual_gw']})")
    print(f"  Model own NA cumulative balance, 2028:     {rec['model_na_cum_balance_2028_gw']}")
    print("\n=== Chips & buildout summary ===")
    print(f"  NVL72 racks 2025/26/27:                    {nv['rack_totals']}")
    print(f"  Chip-implied power GW (PUE {bridge['pue']}):           {bridge['implied_gw']}")
    print(f"  US DC power GW 2023-28 (client):           {client_chips['us_power_gw']}")
    print(f"  Neocloud companies found:                  {len(buildout['neoclouds'])}")
    print(f"  AI labs found:                             {[x['name'] for x in buildout['ai_labs']]}")
    print(f"  US tenant-tag GW: {supply['tenant_gw']}")
    print("\n=== Compute & synthesis summary ===")
    print(f"  Derived TWh (GPU+ASIC) 2023-28:            {energy_split['derived_total_twh']}")
    print(f"  Validation ratio vs client GenAI TWh:      {energy_split['validation_ratio']}")
    print(f"  ASIC parity ExaFLOPs 2024-28:              {compute_split['asic_exa_parity']}")
    print(f"  Constr vs unconstr AI power 2028 (GW):     "
          f"{customer['ai_power_mw'][customer['years'].index(2028)]/1000:.1f} vs "
          f"{unconstrained['ai_power_mw'][unconstrained['years'].index(2028)]/1000:.1f}")
    print(f"  Vintages captured:                         {len(revisions['vintages'])} "
          f"({revisions['vintages'][0]['label']} -> {revisions['vintages'][-1]['label']})")
    print(f"  Slippage: median {slippage['median_mo']} mo, {slippage['pct_ge_7mo']}% >=7mo, n={slippage['n']}")
    print("\n=== Geography & deployment summary ===")
    _i28 = cap_by_deploy["years"].index(2028)
    _cap28 = {k: cap_by_deploy["annual"][k][_i28] for k in ("selfbuilt", "leasing", "neocloud")}
    print(f"  US facilities coming online 2026-28:       {geo['n']} ({geo['total_gw']} GW)")
    print(f"  Map GW by class (self/lease/neo):          {geo['by_class_gw']}")
    print(f"  GW by gas-pipeline proximity (<=5/5-25/>25 mi): {geo['pipe_buckets_gw']}")
    print(f"  Capacity-by-deploy year-end 2028 (GW):     {_cap28}")
    print(f"  Chips-by-deploy 2028 (GW):                 self "
          f"{chips_by_deploy['selfbuilt'][-1]}, lease {chips_by_deploy['leasing'][-1]}, neo {chips_by_deploy['neocloud'][-1]}")
    print("\n=== Chip power & energy demand (client + MS files) ===")
    _fam28 = {f["name"]: f["values"][-1] for f in chip_power_global["families"]}
    print(f"  Global chip DC power GW 2021-28:           {chip_power_global['dc_power_gw']}")
    print(f"  Global chip server power by family 2028:   {_fam28}")
    print(f"  Global chip TWh actual 2021-28:            {chip_power_global['twh_actual']}")
    print(f"  US chip DC power GW 2023-28:               {ai_power_us['dc_power_gw']}")
    print("\n=== Cancelled & delayed (US, original online 2026-2028) ===")
    dc = delayed_cancelled
    print(f"  Mapped facilities:                         {dc['n_facilities']} "
          f"(delayed {dc['n_sites']['delayed']}, cancelled {dc['n_sites']['cancelled']})")
    print(f"  Delayed / cancelled GW:                    {dc['totals_gw']['delayed']} / {dc['totals_gw']['cancelled']}")
    print(f"  Median delay:                              {dc['delay_stats']['median_mo']} mo "
          f"(max {dc['delay_stats']['max_mo']} mo)")
    print(f"  TWh/yr deferred (delayed) / lost (cancelled) @ {dc['intensity_twh_per_gw']} TWh/GW: "
          f"{dc['twh']['delayed_deferred']} / {dc['twh']['cancelled_lost']}")
    print(f"  Externally-cited events / macro reports:   {len(dc['cited_events'])} / {len(dc['cited_macro'])}")
    print(f"\nWrote:\n  {OUT_JSON}\n  {OUT_HTML}\n  {OUT_CHIPS_HTML}\n  {OUT_COMPUTE_HTML}\n  {OUT_SYNTHESIS_HTML}\n  {OUT_GEO_HTML}\n  {OUT_FLOWS_HTML}\n  {OUT_CHIPPOWER_HTML}\n  {OUT_DELAYED_HTML}\n  {OUT_CANCELLEDPROFILE_HTML}\n  {OUT_SHORTFALLMATH_HTML}")


if __name__ == "__main__":
    main()
